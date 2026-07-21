import zipfile
from decimal import Decimal
from io import BytesIO

from apps.common.drive import (
    DriveError,
    DriveFileInfo,
    build_drive_file_tree_map,
    download_drive_file,
    ensure_statement_folder,
    ensure_trash_folder,
    list_drive_folder_contents,
    move_drive_file_to_folder,
    serialize_drive_error,
    upload_file_to_drive,
)
from apps.common.permissions import EditProtectedMixin
from apps.common.services import manage_drive_files
from apps.deals.permissions import build_deal_visibility_q
from django.db import transaction
from django.db.models import (
    Case,
    Count,
    DateField,
    DecimalField,
    F,
    IntegerField,
    OuterRef,
    Prefetch,
    Q,
    Subquery,
    Sum,
    TextField,
    Value,
    When,
)
from django.db.models.functions import Cast, Coalesce, NullIf
from django.http import HttpResponse
from django.utils import timezone
from django.utils.encoding import iri_to_uri
from django.utils.text import get_valid_filename
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .filters import PaymentFilterSet
from .models import FinancialRecord, Payment, Statement
from .permissions import get_deal_from_payment, is_admin_user, user_has_deal_access
from .record_filters import apply_financial_record_filters
from .serializers import (
    FinancialRecordSerializer,
    PaymentSerializer,
    StatementFinancialRecordSerializer,
    StatementSerializer,
)
from .services import build_finance_summary_payload
from .services.statements import (
    ensure_unique_zip_path,
    normalize_statement_amount,
    quantize_money,
    records_with_paid_balance,
    sanitize_drive_filename,
)


class StatementDriveTrashSerializer(serializers.Serializer):
    file_ids = serializers.ListField(
        child=serializers.CharField(),
        min_length=1,
        allow_empty=False,
        required=True,
    )


class StatementDriveDownloadSerializer(serializers.Serializer):
    file_ids = serializers.ListField(
        child=serializers.CharField(),
        min_length=1,
        allow_empty=False,
        required=True,
    )


class StatementMarkPaidSerializer(serializers.Serializer):
    paid_at = serializers.DateField(required=False, allow_null=True)


class StatementApplyAmountSerializer(serializers.Serializer):
    mode = serializers.ChoiceField(choices=("rub", "percent"))
    value = serializers.DecimalField(max_digits=12, decimal_places=4)


class FinancialRecordViewSet(EditProtectedMixin, viewsets.ModelViewSet):
    """ViewSet для финансовых записей (доход/расход)"""

    serializer_class = FinancialRecordSerializer
    ordering_fields = [
        "created_at",
        "updated_at",
        "date",
        "amount",
        "payment_paid_balance",
        "payment_is_paid",
        "payment_sort_date",
        "payment_scheduled_date",
        "payment_scheduled_date_is_null",
        "record_comment_sort",
    ]
    ordering = ["-created_at"]

    def get_serializer_class(self):
        if self.request.query_params.get("statement"):
            return StatementFinancialRecordSerializer
        return super().get_serializer_class()

    def get_queryset(self):
        user = self.request.user
        statement_id = self.request.query_params.get("statement")
        is_statement_list = bool(statement_id)
        is_list = self.action == "list"
        paid_balance_subquery = (
            FinancialRecord.objects.filter(
                payment_id=OuterRef("payment_id"),
                date__isnull=False,
                deleted_at__isnull=True,
            )
            .order_by()
            .values("payment_id")
            .annotate(total=Sum("amount"))
            .values("total")[:1]
        )
        list_related_fields = (
            "payment",
            "payment__policy",
            "payment__policy__client",
            "payment__policy__insured_client",
            "payment__policy__deal",
            "payment__policy__deal__client",
            "payment__policy__insurance_type",
            "payment__policy__sales_channel",
            "payment__deal",
            "payment__deal__client",
        )
        detail_related_fields = (
            "statement",
            *list_related_fields,
        )
        queryset = FinancialRecord.objects.select_related(
            *(list_related_fields if is_list and not is_statement_list else detail_related_fields)
        )
        if is_list and not is_statement_list:
            queryset = queryset.prefetch_related(
                Prefetch(
                    "payment__financial_records",
                    queryset=FinancialRecord.objects.filter(
                        date__isnull=False,
                        deleted_at__isnull=True,
                    )
                    .only("id", "amount", "date", "payment_id")
                    .order_by("-date", "-amount", "id"),
                    to_attr="paid_records",
                )
            )
        queryset = queryset.all().order_by("-date", "-created_at")
        annotations = {
            "payment_is_paid": Case(
                When(payment__actual_date__isnull=False, then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            ),
            "payment_sort_date": Coalesce(
                F("payment__actual_date"),
                F("payment__scheduled_date"),
                output_field=DateField(),
            ),
            "payment_scheduled_date": F("payment__scheduled_date"),
            "payment_scheduled_date_is_null": Case(
                When(payment__scheduled_date__isnull=True, then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            ),
            "record_comment_sort": Coalesce(
                # NOTE: note is TextField, description/source are CharField.
                # Cast to a single type to avoid "mixed types" FieldError in Postgres.
                NullIf(F("note"), Value("", output_field=TextField())),
                NullIf(
                    Cast(F("description"), output_field=TextField()),
                    Value("", output_field=TextField()),
                ),
                NullIf(
                    Cast(F("source"), output_field=TextField()),
                    Value("", output_field=TextField()),
                ),
                Value("", output_field=TextField()),
                output_field=TextField(),
            ),
        }
        annotations["payment_paid_balance"] = Coalesce(
            Subquery(
                paid_balance_subquery,
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            0,
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )
        queryset = queryset.annotate(**annotations)

        # Если пользователь не аутентифицирован, возвращаем все записи (AllowAny режим)
        if not user.is_authenticated:
            return queryset

        # Администраторы видят все финансовые записи
        is_admin = is_admin_user(user)

        if not is_admin:
            # Остальные видят только записи для своих сделок (где user = seller или executor)
            queryset = queryset.filter(
                build_deal_visibility_q(user, prefix="payment__policy__deal__")
                | build_deal_visibility_q(user, prefix="payment__deal__")
            ).distinct()

        queryset = apply_financial_record_filters(queryset, self.request.query_params)

        search_term = (self.request.query_params.get("search") or "").strip()
        if len(search_term) >= 1:
            queryset = queryset.filter(
                Q(payment__policy__number__icontains=search_term)
                | Q(payment__policy__client__name__icontains=search_term)
                | Q(payment__policy__insured_client__name__icontains=search_term)
                | Q(payment__policy__insurance_type__name__icontains=search_term)
                | Q(payment__policy__sales_channel__name__icontains=search_term)
                | Q(payment__policy__deal__title__icontains=search_term)
                | Q(payment__policy__deal__client__name__icontains=search_term)
                | Q(payment__deal__title__icontains=search_term)
                | Q(payment__deal__client__name__icontains=search_term)
                | Q(payment__description__icontains=search_term)
                | Q(description__icontains=search_term)
                | Q(source__icontains=search_term)
                | Q(note__icontains=search_term)
            )

        return queryset

    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request, *args, **kwargs):
        records = self.filter_queryset(self.get_queryset()).iterator(chunk_size=500)

        def format_date(value) -> str:
            if not value:
                return "—"
            if hasattr(value, "strftime"):
                return value.strftime("%d.%m.%Y")
            return str(value)

        def format_money(value) -> str:
            try:
                return f"{float(value):,.2f} ₽".replace(",", " ")
            except Exception:
                return f"{value} ₽"

        workbook = Workbook(write_only=True)
        ws = workbook.create_sheet()
        ws.title = "Финансовые записи"

        headers = [
            "Клиент / сделка",
            "Номер полиса",
            "Тип полиса",
            "Канал продаж",
            "Дата платежа",
            "Платеж, ₽",
            "Сальдо, ₽",
            "Примечание",
            "Сумма, ₽",
        ]
        header_font = Font(bold=True, color="1F2937")
        header_fill = PatternFill("solid", fgColor="F8FAFC")
        wrap_top = Alignment(wrap_text=True, vertical="top")
        currency_number_format = "# ##0.00 [$₽-419]"

        header_cells = []
        for header in headers:
            cell = WriteOnlyCell(ws, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap_top
            header_cells.append(cell)
        ws.append(header_cells)

        ws.freeze_panes = "A2"
        for column, width in {
            "A": 36,
            "B": 18,
            "C": 20,
            "D": 20,
            "E": 16,
            "F": 22,
            "G": 16,
            "H": 32,
            "I": 16,
        }.items():
            ws.column_dimensions[column].width = width

        for record in records:
            payment = getattr(record, "payment", None)
            policy = getattr(payment, "policy", None) if payment else None
            deal = get_deal_from_payment(payment)
            statement = getattr(record, "statement", None)

            deal_title = getattr(deal, "title", None) or "-"
            deal_client_name = (
                getattr(getattr(deal, "client", None), "name", None) or "-"
            )
            policy_number = (
                getattr(policy, "number", None)
                or getattr(payment, "policy_number", None)
                or "-"
            )
            policy_type = (
                getattr(getattr(policy, "insurance_type", None), "name", None) or "-"
            )
            sales_channel = (
                getattr(getattr(policy, "sales_channel", None), "name", None) or "-"
            )
            policy_client_name = (
                getattr(getattr(policy, "client", None), "name", None)
                or getattr(getattr(policy, "insured_client", None), "name", None)
                or deal_client_name
                or "-"
            )

            payment_amount = getattr(payment, "amount", None)
            payment_actual = getattr(payment, "actual_date", None)
            payment_scheduled = getattr(payment, "scheduled_date", None)
            payment_cell = (
                f"{format_money(payment_amount)}\n"
                + (
                    f"Оплачен: {format_date(payment_actual)}"
                    if payment_actual
                    else f"Не оплачен (план: {format_date(payment_scheduled)})"
                )
                if payment_amount is not None
                else "—"
            )

            paid_records = getattr(payment, "paid_records", []) if payment else []
            saldo_value = (
                sum((paid_record.amount for paid_record in paid_records), 0)
                if paid_records
                else 0
            )
            comment_parts = [
                (record.note or "").strip(),
                (record.description or "").strip(),
                (record.source or "").strip(),
            ]
            comment_parts = [part for part in comment_parts if part]
            comment_cell = comment_parts[0] if comment_parts else "—"
            if len(comment_parts) > 1:
                comment_cell = f"{comment_cell}\n" + " · ".join(comment_parts[1:])
            if statement:
                if statement.paid_at:
                    comment_cell += f"\nВедомость от {format_date(statement.paid_at)}: {statement.name}"
                else:
                    comment_cell += f"\nВедомость: {statement.name}"

            client_cell = f"{policy_client_name}\n{deal_title}\nКонтакт по сделке: {deal_client_name}"
            values = [
                client_cell,
                policy_number,
                policy_type,
                sales_channel,
                format_date(payment_scheduled),
                payment_cell,
                saldo_value,
                comment_cell,
                abs(record.amount),
            ]
            cells = []
            for index, value in enumerate(values):
                cell = WriteOnlyCell(ws, value=value)
                cell.alignment = wrap_top
                if index in {6, 8}:
                    cell.number_format = currency_number_format
                cells.append(cell)
            ws.append(cells)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        now = timezone.localtime(timezone.now())
        filename = f"financial_records_{now.strftime('%d_%m_%Y_%H_%M_%S')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{iri_to_uri(filename)}"
        )
        return response

    def _can_modify(self, user, instance):
        payment = getattr(instance, "payment", None)
        deal = get_deal_from_payment(payment)
        return user_has_deal_access(user, deal, allow_executor=False)

    def perform_create(self, serializer):
        payment = serializer.validated_data.get("payment")
        deal = get_deal_from_payment(payment)
        if not user_has_deal_access(self.request.user, deal, allow_executor=False):
            raise PermissionDenied("Нет доступа к платежу или сделке.")
        serializer.save()

    def perform_update(self, serializer):
        instance = serializer.instance
        statement = getattr(instance, "statement", None)
        if statement and statement.paid_at:
            raise ValidationError("Нельзя изменять записи в выплаченной ведомости.")
        super().perform_update(serializer)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        statement = getattr(instance, "statement", None)
        if statement:
            if statement.paid_at:
                raise ValidationError("Нельзя удалить запись из выплаченной ведомости.")
            # Если запись уже добавлена в черновик ведомости, удалять её нельзя.
            # Сначала нужно убрать запись из ведомости, чтобы не нарушить её состав.
            raise ValidationError(
                "Нельзя удалить запись из ведомости. Сначала уберите её из состава ведомости."
            )
        return super().destroy(request, *args, **kwargs)

    def get_object(self):
        from django.shortcuts import get_object_or_404

        kwargs = {self.lookup_field: self.kwargs.get(self.lookup_field)}
        lookup = FinancialRecord.objects.select_related("payment")
        return get_object_or_404(lookup, **kwargs)


class StatementViewSet(EditProtectedMixin, viewsets.ModelViewSet):
    serializer_class = StatementSerializer
    ordering_fields = [
        "created_at",
        "updated_at",
        "paid_at",
        "status",
        "statement_type",
    ]
    ordering = ["-created_at"]
    owner_field = "created_by"

    def get_queryset(self):
        user = self.request.user
        active_records = FinancialRecord.objects.filter(
            statement=OuterRef("pk"),
            deleted_at__isnull=True,
        ).values("statement")
        queryset = (
            Statement.objects.annotate(
                records_count=Coalesce(
                    Subquery(
                        active_records.annotate(total=Count("id")).values("total")[:1],
                        output_field=IntegerField(),
                    ),
                    Value(0),
                    output_field=IntegerField(),
                ),
                total_amount=Coalesce(
                    Subquery(
                        active_records.annotate(total=Sum("amount")).values("total")[
                            :1
                        ],
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    ),
                    Value(0),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                ),
            )
            .all()
            .order_by("-created_at")
        )
        if not user.is_authenticated:
            return queryset.none()
        if is_admin_user(user):
            return queryset
        return queryset.filter(
            Q(created_by=user)
            | build_deal_visibility_q(user, prefix="records__payment__deal__")
            | build_deal_visibility_q(user, prefix="records__payment__policy__deal__")
        ).distinct()

    def perform_create(self, serializer):
        record_ids = serializer.validated_data.get("record_ids") or []
        self._validate_record_access(record_ids)
        serializer.save(created_by=self.request.user)

    def _validate_record_access(self, records):
        for record in records:
            deal = get_deal_from_payment(getattr(record, "payment", None))
            if not user_has_deal_access(self.request.user, deal, allow_executor=False):
                raise PermissionDenied("Нет доступа к финансовой записи для ведомости.")

    @staticmethod
    def _parse_record_ids(data):
        record_ids = data.get("record_ids") or []
        if not isinstance(record_ids, list) or not record_ids:
            raise ValidationError(
                {"record_ids": "Укажите непустой список идентификаторов."}
            )
        return list(dict.fromkeys(record_ids))

    @staticmethod
    def _set_statement_summary(statement):
        summary = FinancialRecord.objects.filter(
            statement=statement,
            deleted_at__isnull=True,
        ).aggregate(records_count=Count("id"), total_amount=Sum("amount"))
        statement.records_count = summary["records_count"] or 0
        statement.total_amount = summary["total_amount"] or Decimal("0")
        return statement

    def _locked_statement(self, statement_id):
        statement = Statement.objects.select_for_update().get(pk=statement_id)
        if not self._can_modify(self.request.user, statement):
            raise PermissionDenied(
                "Только администратор или владелец может изменять ведомость."
            )
        if statement.paid_at:
            raise ValidationError("Нельзя изменять выплаченную ведомость.")
        return statement

    def _locked_records(self, record_ids):
        records = list(
            FinancialRecord.objects.filter(id__in=record_ids, deleted_at__isnull=True)
            .select_for_update(of=("self",))
            .select_related(
                "payment",
                "payment__policy",
                "payment__policy__deal",
                "payment__deal",
            )
            .order_by("pk")
        )
        if len(records) != len(record_ids):
            raise ValidationError(
                {"record_ids": "Одна или несколько записей не найдены."}
            )
        self._validate_record_access(records)
        return records

    @action(detail=True, methods=["post"], url_path="attach-records")
    def attach_records(self, request, *args, **kwargs):
        record_ids = self._parse_record_ids(request.data)
        with transaction.atomic():
            statement = self._locked_statement(self.kwargs[self.lookup_field])
            records = self._locked_records(record_ids)
            wrong_type = [
                record.id
                for record in records
                if record.record_type
                != (
                    FinancialRecord.RecordType.INCOME
                    if statement.statement_type == Statement.TYPE_INCOME
                    else FinancialRecord.RecordType.EXPENSE
                )
            ]
            if wrong_type:
                raise ValidationError(
                    {"record_ids": "Тип записей не соответствует типу ведомости."}
                )
            conflicting_ids = [
                record.id
                for record in records
                if record.statement_id and record.statement_id != statement.id
            ]
            if conflicting_ids:
                raise ValidationError(
                    {"record_ids": "Запись уже включена в другую ведомость."}
                )
            attached_ids = [record.id for record in records if not record.statement_id]
            if attached_ids:
                FinancialRecord.objects.filter(id__in=attached_ids).update(
                    statement=statement
                )
                statement.updated_at = timezone.now()
                statement.save(update_fields=["updated_at"])
            self._set_statement_summary(statement)

        return Response(
            {
                "attached_count": len(attached_ids),
                "attached_record_ids": [str(record_id) for record_id in attached_ids],
                "statement": StatementSerializer(
                    statement, context={"request": request}
                ).data,
            }
        )

    def _create_download_response(
        self,
        content: bytes,
        filename: str,
        content_type: str = "application/octet-stream",
    ) -> HttpResponse:
        response = HttpResponse(content, content_type=content_type)
        raw_name = (filename or "").strip() or "download"
        safe_name = get_valid_filename(raw_name) or "download"
        response["Content-Disposition"] = (
            f"attachment; filename=\"{safe_name}\"; filename*=UTF-8''{iri_to_uri(raw_name)}"
        )
        return response

    def perform_update(self, serializer):
        record_ids = serializer.validated_data.get("record_ids") or []
        if record_ids:
            self._validate_record_access(record_ids)
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.paid_at:
            raise ValidationError("Нельзя удалять выплаченную ведомость.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=["post"], url_path="remove-records")
    def remove_records(self, request, *args, **kwargs):
        record_ids = self._parse_record_ids(request.data)
        with transaction.atomic():
            statement = self._locked_statement(self.kwargs[self.lookup_field])
            records = self._locked_records(record_ids)
            if any(record.statement_id != statement.id for record in records):
                raise ValidationError(
                    {"record_ids": "Запись не входит в эту ведомость."}
                )
            FinancialRecord.objects.filter(id__in=record_ids).update(statement=None)
            statement.updated_at = timezone.now()
            statement.save(update_fields=["updated_at"])
            self._set_statement_summary(statement)

        return Response(
            {
                "removed": len(record_ids),
                "removed_count": len(record_ids),
                "removed_record_ids": [str(record_id) for record_id in record_ids],
                "statement": StatementSerializer(
                    statement, context={"request": request}
                ).data,
            }
        )

    @action(detail=True, methods=["post"], url_path="apply-amount")
    def apply_amount(self, request, *args, **kwargs):
        statement = self.get_object()
        if statement.paid_at:
            raise ValidationError("Нельзя изменять выплаченную ведомость.")

        serializer = StatementApplyAmountSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        mode = serializer.validated_data["mode"]
        value = abs(serializer.validated_data["value"])

        records = list(records_with_paid_balance(statement))
        self._validate_record_access(records)

        updated_records = []
        unchanged = 0
        skipped_reasons = {"zero_balance": 0}

        for record in records:
            if mode == "percent":
                base = abs(getattr(record, "payment_paid_balance", Decimal("0")) or 0)
                if base <= 0:
                    skipped_reasons["zero_balance"] += 1
                    continue
                absolute_amount = quantize_money((base * value) / Decimal("100"))
            else:
                absolute_amount = quantize_money(value)

            next_amount = normalize_statement_amount(record, absolute_amount)
            if quantize_money(record.amount) == next_amount:
                unchanged += 1
                continue

            record.amount = next_amount
            record.updated_at = timezone.now()
            updated_records.append(record)

        with transaction.atomic():
            if updated_records:
                FinancialRecord.objects.bulk_update(
                    updated_records, ["amount", "updated_at"]
                )
                statement.updated_at = timezone.now()
                statement.save(update_fields=["updated_at"])

        response_records = list(records_with_paid_balance(statement))
        payload = {
            "updated": len(updated_records),
            "unchanged": unchanged,
            "skipped": sum(skipped_reasons.values()),
            "skipped_reasons": skipped_reasons,
            "records": FinancialRecordSerializer(
                response_records, many=True, context={"request": request}
            ).data,
            "statement": StatementSerializer(
                statement, context={"request": request}
            ).data,
        }
        return Response(payload)

    @action(detail=True, methods=["post"], url_path="mark-paid")
    def mark_paid(self, request, *args, **kwargs):
        statement = self.get_object()
        if statement.paid_at:
            raise ValidationError("У ведомости уже указана дата выплаты.")

        serializer = StatementMarkPaidSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        paid_at = serializer.validated_data.get("paid_at") or statement.paid_at
        if not paid_at:
            raise ValidationError({"paid_at": "Укажите дату оплаты ведомости."})

        with transaction.atomic():
            if statement.paid_at != paid_at:
                statement.paid_at = paid_at
                # Keep status consistent for admin/UI legacy, but business logic relies on paid_at.
                statement.status = Statement.STATUS_PAID
                statement.save(update_fields=["paid_at", "status", "updated_at"])

            FinancialRecord.objects.filter(
                statement=statement, deleted_at__isnull=True
            ).update(date=paid_at)

        payload = StatementSerializer(statement, context={"request": request}).data
        return Response(payload)

    @action(detail=True, methods=["post"], url_path="export-xlsx")
    def export_xlsx(self, request, *args, **kwargs):
        statement = self.get_object()

        try:
            folder_id = statement.drive_folder_id or ensure_statement_folder(statement)
        except DriveError as exc:
            raise ValidationError({"detail": str(exc)}) from exc

        if not folder_id:
            raise ValidationError(
                {"detail": "Папка Google Drive для ведомости не найдена."}
            )

        now = timezone.localtime(timezone.now())
        ts = now.strftime("%d_%m_%Y_%H_%M_%S")
        base_name = sanitize_drive_filename(statement.name or "Ведомость")
        filename = f"{base_name}_{ts}.xlsx"

        records_qs = (
            FinancialRecord.objects.filter(statement=statement, deleted_at__isnull=True)
            .select_related(
                "payment",
                "payment__policy",
                "payment__policy__deal",
                "payment__policy__deal__client",
                "payment__policy__insurance_type",
                "payment__policy__sales_channel",
                "payment__deal",
                "payment__deal__client",
            )
            .order_by("created_at", "id")
        )
        records = list(records_qs)

        payment_ids = {record.payment_id for record in records if record.payment_id}
        payments = (
            Payment.objects.filter(id__in=payment_ids)
            .select_related(
                "policy",
                "policy__deal",
                "policy__deal__client",
                "policy__insurance_type",
                "policy__sales_channel",
                "deal",
                "deal__client",
            )
            .prefetch_related(
                Prefetch(
                    "financial_records",
                    queryset=FinancialRecord.objects.filter(
                        date__isnull=False, deleted_at__isnull=True
                    )
                    .only("id", "amount", "date", "payment_id")
                    .order_by("-date", "-amount", "id"),
                    to_attr="paid_records",
                )
            )
        )
        payments_by_id = {payment.id: payment for payment in payments}

        def format_date(value) -> str:
            if not value:
                return "—"
            if hasattr(value, "strftime"):
                return value.strftime("%d.%m.%Y")
            return str(value)

        def format_money(value) -> str:
            try:
                return f"{float(value):,.2f} ₽".replace(",", " ")
            except Exception:
                return f"{value} ₽"

        workbook = Workbook()
        ws = workbook.active
        ws.title = "Ведомость"

        headers = [
            "Клиент / сделка",
            "Номер полиса",
            "Тип полиса",
            "Канал продаж",
            "Дата платежа",
            "Платеж, ₽",
            "Доходы / расходы",
            "Сальдо, ₽",
            "Примечание",
            "Сумма, ₽",
        ]

        header_font = Font(bold=True, color="1F2937")
        header_fill = PatternFill("solid", fgColor="F8FAFC")
        wrap_top = Alignment(wrap_text=True, vertical="top")
        currency_number_format = "# ##0.00 [$₽-419]"

        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap_top

        ws.freeze_panes = "A2"

        # Column widths are approximate; UI has horizontal scroll anyway.
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 22
        ws.column_dimensions["G"].width = 28
        ws.column_dimensions["H"].width = 16
        ws.column_dimensions["I"].width = 28
        ws.column_dimensions["J"].width = 16

        for record in records:
            payment = payments_by_id.get(record.payment_id)
            policy = getattr(payment, "policy", None) if payment else None
            deal = getattr(payment, "deal", None) if payment else None

            deal_title = getattr(deal, "title", None) or "-"
            deal_client_name = (
                getattr(getattr(deal, "client", None), "name", None) or "-"
            )
            policy_number = (
                getattr(policy, "number", None)
                or getattr(payment, "policy_number", None)
                or "-"
            )
            policy_type = (
                getattr(getattr(policy, "insurance_type", None), "name", None) or "-"
            )
            sales_channel = (
                getattr(getattr(policy, "sales_channel", None), "name", None) or "-"
            )

            policy_client_name = (
                getattr(getattr(policy, "client", None), "name", None)
                or getattr(getattr(policy, "insured_client", None), "name", None)
                or deal_client_name
                or "-"
            )

            payment_amount = getattr(payment, "amount", None)
            payment_actual = getattr(payment, "actual_date", None)
            payment_scheduled = getattr(payment, "scheduled_date", None)
            payment_cell = (
                f"{format_money(payment_amount)}\n"
                + (
                    f"Оплачен: {format_date(payment_actual)}"
                    if payment_actual
                    else f"Не оплачен (план: {format_date(payment_scheduled)})"
                )
                if payment_amount is not None
                else "—"
            )

            paid_records = getattr(payment, "paid_records", []) if payment else []
            saldo_value = (
                sum((pr.amount for pr in paid_records), 0) if paid_records else 0
            )
            if paid_records:
                operations_lines = []
                for pr in paid_records:
                    entry_type = "Доход" if pr.amount >= 0 else "Расход"
                    operations_lines.append(
                        f"{entry_type} {format_money(abs(pr.amount))} · {format_date(pr.date)}"
                    )
                operations_cell = "\n".join(operations_lines)
            else:
                operations_cell = "Операций нет"

            parts = [
                (record.note or "").strip(),
                (record.description or "").strip(),
                (record.source or "").strip(),
            ]
            parts = [p for p in parts if p]
            comment_cell = parts[0] if parts else "—"
            if len(parts) > 1:
                comment_cell = f"{comment_cell}\n" + " · ".join(parts[1:])

            if record.statement_id:
                if statement.paid_at:
                    comment_cell += f"\nВедомость от {format_date(statement.paid_at)}: {statement.name}"
                else:
                    comment_cell += f"\nВедомость: {statement.name}"

            amount_value = abs(record.amount)

            client_cell = f"{policy_client_name}\n{deal_title}\nКонтакт по сделке: {deal_client_name}"

            ws.append(
                [
                    client_cell,
                    policy_number,
                    policy_type,
                    sales_channel,
                    format_date(payment_scheduled),
                    payment_cell,
                    operations_cell,
                    saldo_value,
                    comment_cell,
                    amount_value,
                ]
            )

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap_top
            row[7].number_format = currency_number_format
            row[9].number_format = currency_number_format

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        try:
            drive_file: DriveFileInfo = upload_file_to_drive(
                folder_id,
                buffer,
                filename,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except DriveError as exc:
            raise ValidationError({"detail": str(exc)}) from exc

        return Response({"folder_id": folder_id, "file": drive_file})

    @action(
        detail=True,
        methods=["get", "post", "delete"],
        url_path="drive-files",
        parser_classes=[MultiPartParser, FormParser, JSONParser],
    )
    def drive_files(self, request, *args, **kwargs):
        statement = self.get_object()
        uploaded_file = request.FILES.get("file") if request.method == "POST" else None

        if request.method == "DELETE":
            serializer = StatementDriveTrashSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            file_ids = [
                file_id.strip()
                for file_id in serializer.validated_data["file_ids"]
                if isinstance(file_id, str) and file_id.strip()
            ]
            if not file_ids:
                raise ValidationError({"file_ids": "Нужно передать ID файлов."})

            try:
                folder_id = statement.drive_folder_id or ensure_statement_folder(
                    statement
                )
                if not folder_id:
                    return Response(
                        {"detail": "Папка Google Drive для ведомости не найдена."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                drive_files = list_drive_folder_contents(folder_id)
                drive_file_map = {item["id"]: item for item in drive_files}
                missing_file_ids = [
                    file_id
                    for file_id in file_ids
                    if file_id not in drive_file_map
                    or drive_file_map[file_id]["is_folder"]
                ]
                if missing_file_ids:
                    return Response(
                        {
                            "detail": "Файлы не найдены или это папки.",
                            "missing_file_ids": missing_file_ids,
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                trash_folder_id = ensure_trash_folder(folder_id)
                for file_id in file_ids:
                    move_drive_file_to_folder(file_id, trash_folder_id)

                return Response(
                    {
                        "moved_file_ids": file_ids,
                        "trash_folder_id": trash_folder_id,
                    }
                )
            except DriveError as exc:
                return Response(
                    serialize_drive_error(exc),
                    status=status.HTTP_503_SERVICE_UNAVAILABLE,
                )

        if request.method == "POST" and not uploaded_file:
            return Response(
                {"detail": "No file provided for upload."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            result = manage_drive_files(
                instance=statement,
                ensure_folder_func=ensure_statement_folder,
                uploaded_file=uploaded_file,
            )
            return Response(result)
        except DriveError as exc:
            return Response(
                serialize_drive_error(exc),
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

    @action(
        detail=True,
        methods=["post"],
        url_path="drive-files/download",
        parser_classes=[JSONParser],
    )
    def download_drive_files(self, request, *args, **kwargs):
        statement = self.get_object()
        serializer = StatementDriveDownloadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        file_ids = [
            file_id.strip()
            for file_id in serializer.validated_data["file_ids"]
            if isinstance(file_id, str) and file_id.strip()
        ]
        if not file_ids:
            raise ValidationError({"file_ids": "Нужно передать ID файлов."})

        try:
            folder_id = statement.drive_folder_id or ensure_statement_folder(statement)
            if not folder_id:
                return Response(
                    {"detail": "Папка Google Drive для ведомости не найдена."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            drive_file_map: dict[str, DriveFileInfo] = build_drive_file_tree_map(
                folder_id
            )
            missing_ids = [
                file_id for file_id in file_ids if file_id not in drive_file_map
            ]
            if missing_ids:
                return Response(
                    {
                        "detail": "Файлы не найдены в папке ведомости.",
                        "missing_file_ids": missing_ids,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            selected_items = [drive_file_map[file_id] for file_id in file_ids]
            folder_ids = [item["id"] for item in selected_items if item["is_folder"]]
            if folder_ids:
                return Response(
                    {
                        "detail": "Скачивание папок не поддерживается. Выберите только файлы."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if len(selected_items) == 1:
                item = selected_items[0]
                content = download_drive_file(item["id"])
                return self._create_download_response(
                    content=content,
                    filename=item["name"],
                    content_type=item.get("mime_type") or "application/octet-stream",
                )

            zip_buffer = BytesIO()
            with zipfile.ZipFile(
                zip_buffer, "w", compression=zipfile.ZIP_DEFLATED
            ) as zip_file:
                seen_paths: set[str] = set()
                for item in selected_items:
                    content = download_drive_file(item["id"])
                    zip_path = ensure_unique_zip_path(
                        item["name"] or "file", seen_paths
                    )
                    zip_file.writestr(zip_path, content)

            zip_buffer.seek(0)
            archive_name = f"{statement.name or f'statement-{statement.id}'}-files.zip"
            return self._create_download_response(
                content=zip_buffer.read(),
                filename=archive_name,
                content_type="application/zip",
            )
        except DriveError as exc:
            return Response(
                serialize_drive_error(exc),
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )


class PaymentViewSet(EditProtectedMixin, viewsets.ModelViewSet):
    """ViewSet для платежей с поддержкой проверки удаления"""

    serializer_class = PaymentSerializer
    filterset_class = PaymentFilterSet
    search_fields = ["description", "policy__deal__title", "deal__title"]
    ordering_fields = [
        "created_at",
        "updated_at",
        "scheduled_date",
        "actual_date",
        "amount",
    ]
    ordering = ["-created_at"]

    def get_queryset(self):
        user = self.request.user
        queryset = (
            Payment.objects.select_related("policy", "policy__deal", "deal")
            .prefetch_related("financial_records")
            .all()
            .order_by("-scheduled_date")
        )

        # Если пользователь не аутентифицирован, возвращаем все записи (AllowAny режим)
        if not user.is_authenticated:
            return queryset

        # Администраторы видят все платежи
        is_admin = is_admin_user(user)

        if not is_admin:
            # Остальные видят только платежи для своих сделок (где user = seller или executor)
            queryset = queryset.filter(
                build_deal_visibility_q(user, prefix="policy__deal__")
                | build_deal_visibility_q(user, prefix="deal__")
            ).distinct()

        return queryset

    def perform_create(self, serializer):
        policy = serializer.validated_data.get("policy")
        deal = getattr(policy, "deal", None)
        if not user_has_deal_access(self.request.user, deal):
            raise PermissionDenied("Нет доступа к сделке.")
        serializer.save()

    def _can_modify(self, user, instance):
        deal = get_deal_from_payment(instance)
        return user_has_deal_access(user, deal)

    def destroy(self, request, *args, **kwargs):
        """Delete payment only if it has no paid records and is unpaid."""
        instance = self.get_object()

        if instance.financial_records.filter(
            date__isnull=False, deleted_at__isnull=True
        ).exists():
            return Response(
                {
                    "detail": "Нельзя удалить платёж, пока у него есть оплаченные финансовые записи."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not instance.can_delete():
            return Response(
                {"detail": "Нельзя удалить оплаченный платёж."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return super().destroy(request, *args, **kwargs)


class FinanceSummaryView(APIView):
    """Endpoint для сводки по финансам"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        payload = build_finance_summary_payload(user=request.user)
        return Response(payload)
