import zipfile
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from apps.clients.models import Client
from apps.common.drive import ensure_statement_folder
from apps.common.tests.auth_utils import AuthenticatedAPITestCase
from apps.deals.models import Deal, InsuranceCompany, InsuranceType, SalesChannel
from apps.finances.models import FinancialRecord, Payment, Statement
from apps.policies.models import Policy
from apps.tasks.models import Task
from apps.users.models import Role, UserRole
from django.contrib.auth.models import User
from django.db import connection
from django.test.utils import CaptureQueriesContext
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status


class FinanceAccessTests(AuthenticatedAPITestCase):
    """Проверяем доступ покупателей к платежам и финансовым записям сделок."""

    def setUp(self):
        super().setUp()
        self.seller = User.objects.create_user(  # pragma: allowlist secret
            username="seller", password="pass"  # pragma: allowlist secret
        )
        self.executor = User.objects.create_user(  # pragma: allowlist secret
            username="executor", password="pass"  # pragma: allowlist secret
        )
        self.visible_user = User.objects.create_user(  # pragma: allowlist secret
            username="viewer", password="pass"  # pragma: allowlist secret
        )
        self.task_assignee = User.objects.create_user(  # pragma: allowlist secret
            username="tasker", password="pass"  # pragma: allowlist secret
        )
        self.other_user = User.objects.create_user(  # pragma: allowlist secret
            username="other", password="pass"  # pragma: allowlist secret
        )
        self.admin_user = User.objects.create_user(  # pragma: allowlist secret
            username="admin", password="pass"  # pragma: allowlist secret
        )
        self.superuser = User.objects.create_superuser(  # pragma: allowlist secret
            username="root",
            password="pass",  # pragma: allowlist secret
            email="root@example.com",  # pragma: allowlist secret
        )
        self.localized_admin = User.objects.create_user(  # pragma: allowlist secret
            username="localized-admin", password="pass"  # pragma: allowlist secret
        )

        client = Client.objects.create(name="Client")
        self.deal = Deal.objects.create(
            title="Permission Deal",
            client=client,
            seller=self.seller,
            executor=self.executor,
            status="open",
            stage_name="initial",
        )
        self.deal.visible_users.add(self.visible_user)
        Task.objects.create(
            title="Finance follow-up",
            deal=self.deal,
            assignee=self.task_assignee,
        )
        self.insurance_company = InsuranceCompany.objects.create(name="Finance IC")
        self.insurance_type = InsuranceType.objects.create(name="Finance Type")
        self.policy = Policy.objects.create(
            number="FIN-POL-001",
            deal=self.deal,
            client=client,
            insurance_company=self.insurance_company,
            insurance_type=self.insurance_type,
        )

        admin_role, _ = Role.objects.get_or_create(
            name="Admin", defaults={"description": "Системный администратор"}
        )
        UserRole.objects.create(user=self.admin_user, role=admin_role)
        localized_admin_role, _ = Role.objects.get_or_create(
            name="Администратор",
            defaults={"description": "Локализованный администратор"},
        )
        UserRole.objects.create(user=self.localized_admin, role=localized_admin_role)

        self.payment = Payment.objects.create(
            policy=self.policy,
            deal=self.deal,
            amount=Decimal("1000.00"),
            description="Initial",
        )
        self.fin_record = FinancialRecord.objects.create(
            payment=self.payment, amount=Decimal("100"), description="Calc"
        )

    def _extract_results(self, response):
        payload = response.json()
        return payload.get("results", payload)

    def test_seller_can_create_payment(self):
        self.authenticate(self.seller)
        response = self.api_client.post(
            "/api/v1/payments/",
            {
                "amount": "1500.00",
                "policy": str(self.policy.id),
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_executor_can_create_payment(self):
        self.authenticate(self.executor)
        response = self.api_client.post(
            "/api/v1/payments/",
            {
                "amount": "2000.00",
                "policy": str(self.policy.id),
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_other_user_cannot_create_payment(self):
        self.authenticate(self.other_user)
        response = self.api_client.post(
            "/api/v1/payments/",
            {
                "amount": "2000.00",
                "policy": str(self.policy.id),
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_seller_can_create_financial_record(self):
        self.authenticate(self.seller)
        response = self.api_client.post(
            "/api/v1/financial_records/",
            {
                "payment": str(self.payment.id),
                "amount": "50.00",
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_create_expense_with_positive_amount_is_normalized_by_record_type(self):
        self.authenticate(self.seller)
        response = self.api_client.post(
            "/api/v1/financial_records/",
            {
                "payment": str(self.payment.id),
                "amount": "50.00",
                "record_type": "expense",
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        payload = response.json()
        self.assertEqual(payload["record_type"], "Расход")

        created_record = FinancialRecord.objects.get(id=payload["id"])
        self.assertEqual(created_record.record_type, FinancialRecord.RecordType.EXPENSE)
        self.assertEqual(created_record.amount, Decimal("-50.00"))

    def test_executor_cannot_create_financial_record(self):
        self.authenticate(self.executor)
        response = self.api_client.post(
            "/api/v1/financial_records/",
            {
                "payment": str(self.payment.id),
                "amount": "75.00",
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_other_user_cannot_create_financial_record(self):
        self.authenticate(self.other_user)
        response = self.api_client.post(
            "/api/v1/financial_records/",
            {
                "payment": str(self.payment.id),
                "amount": "75.00",
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_visible_user_can_list_payments_for_visible_deal(self):
        self.authenticate(self.visible_user)

        response = self.api_client.get("/api/v1/payments/", {"deal": str(self.deal.id)})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        self.assertEqual([item["id"] for item in results], [str(self.payment.id)])

    def test_task_assignee_can_list_payments_for_related_deal(self):
        self.authenticate(self.task_assignee)

        response = self.api_client.get("/api/v1/payments/", {"deal": str(self.deal.id)})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        self.assertEqual([item["id"] for item in results], [str(self.payment.id)])

    def test_superuser_can_list_all_financial_records_without_admin_role(self):
        self.authenticate(self.superuser)

        response = self.api_client.get("/api/v1/financial_records/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        self.assertEqual([item["id"] for item in results], [str(self.fin_record.id)])

    def test_localized_admin_role_is_treated_as_admin(self):
        self.authenticate(self.localized_admin)

        response = self.api_client.get("/api/v1/financial_records/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        self.assertEqual([item["id"] for item in results], [str(self.fin_record.id)])

    def test_deal_filter_on_payments_includes_policy_linked_payments(self):
        policy = Policy.objects.create(number="POLICY-DEAL-FILTER", deal=self.deal)
        policy_payment = Payment.objects.create(
            policy=policy,
            amount=Decimal("800.00"),
            description="Policy-linked payment",
        )

        self.authenticate(self.visible_user)
        response = self.api_client.get("/api/v1/payments/", {"deal": str(self.deal.id)})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        returned_ids = {item["id"] for item in results}
        self.assertIn(str(self.payment.id), returned_ids)
        self.assertIn(str(policy_payment.id), returned_ids)

    def test_visible_user_can_list_financial_records_for_visible_deal(self):
        self.authenticate(self.visible_user)

        response = self.api_client.get(
            "/api/v1/financial_records/",
            {"deal": str(self.deal.id)},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        self.assertEqual([item["id"] for item in results], [str(self.fin_record.id)])

    def test_task_assignee_can_list_financial_records_for_related_deal(self):
        self.authenticate(self.task_assignee)

        response = self.api_client.get(
            "/api/v1/financial_records/",
            {"deal": str(self.deal.id)},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        self.assertEqual([item["id"] for item in results], [str(self.fin_record.id)])

    def test_seller_can_update_financial_record(self):
        self.authenticate(self.seller)
        response = self.api_client.patch(
            f"/api/v1/financial_records/{self.fin_record.id}/",
            {"amount": "-50.00"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_update_expense_with_positive_amount_keeps_negative_sign(self):
        self.fin_record.record_type = FinancialRecord.RecordType.EXPENSE
        self.fin_record.amount = Decimal("-100.00")
        self.fin_record.save()

        self.authenticate(self.seller)
        response = self.api_client.patch(
            f"/api/v1/financial_records/{self.fin_record.id}/",
            {"amount": "75.00", "record_type": "expense"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        self.assertEqual(payload["record_type"], "Расход")

        self.fin_record.refresh_from_db()
        self.assertEqual(
            self.fin_record.record_type, FinancialRecord.RecordType.EXPENSE
        )
        self.assertEqual(self.fin_record.amount, Decimal("-75.00"))

    def test_update_legacy_expense_with_missing_record_type_keeps_expense_type(self):
        FinancialRecord.objects.filter(id=self.fin_record.id).update(
            record_type="Расход",
            amount=Decimal("-100.00"),
        )
        self.fin_record.refresh_from_db()

        self.authenticate(self.seller)
        response = self.api_client.patch(
            f"/api/v1/financial_records/{self.fin_record.id}/",
            {"amount": "75.00"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        self.assertEqual(payload["record_type"], "Расход")

        self.fin_record.refresh_from_db()
        self.assertEqual(
            self.fin_record.record_type, FinancialRecord.RecordType.EXPENSE
        )
        self.assertEqual(self.fin_record.amount, Decimal("-75.00"))

    def test_financial_record_list_includes_enriched_payment_fields(self):
        self.authenticate(self.seller)

        client = self.deal.client
        client.name = "Мария"
        client.save(update_fields=["name"])
        self.payment.description = "Комиссия"
        self.payment.actual_date = timezone.now().date()
        self.payment.scheduled_date = timezone.now().date()
        self.payment.save(
            update_fields=["description", "actual_date", "scheduled_date"]
        )

        response = self.api_client.get("/api/v1/financial_records/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        self.assertTrue(results)
        record = results[0]
        self.assertEqual(record["deal_id"], str(self.deal.id))
        self.assertEqual(record["deal_title"], self.deal.title)
        self.assertEqual(record["deal_client_name"], "Мария")
        self.assertEqual(record["payment_description"], "Комиссия")
        self.assertEqual(
            record["payment_actual_date"], self.payment.actual_date.isoformat()
        )
        self.assertEqual(
            record["payment_scheduled_date"], self.payment.scheduled_date.isoformat()
        )
        self.assertEqual(record["policy_id"], str(self.policy.id))
        self.assertEqual(record["policy_number"], self.policy.number)
        self.assertIsNone(record["sales_channel_name"])

    def test_financial_record_search_still_works_with_enriched_fields(self):
        self.authenticate(self.seller)
        self.fin_record.note = "УникальныйПоиск123"
        self.fin_record.save(update_fields=["note"])

        response = self.api_client.get(
            "/api/v1/financial_records/",
            {"search": "УникальныйПоиск123"},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]["id"], str(self.fin_record.id))

    def test_executor_cannot_update_financial_record(self):
        self.authenticate(self.executor)
        response = self.api_client.patch(
            f"/api/v1/financial_records/{self.fin_record.id}/",
            {"amount": "-25.00"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_other_user_cannot_update_financial_record(self):
        self.authenticate(self.other_user)
        response = self.api_client.patch(
            f"/api/v1/financial_records/{self.fin_record.id}/",
            {"amount": "25.00"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_executor_cannot_delete_financial_record(self):
        self.authenticate(self.executor)
        response = self.api_client.delete(
            f"/api/v1/financial_records/{self.fin_record.id}/"
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_other_user_cannot_delete_financial_record(self):
        self.authenticate(self.other_user)
        response = self.api_client.delete(
            f"/api/v1/financial_records/{self.fin_record.id}/"
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_seller_can_delete_unpaid_payment(self):
        self.authenticate(self.seller)
        response = self.api_client.delete(f"/api/v1/payments/{self.payment.id}/")
        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)

    def test_cannot_delete_paid_payment(self):
        paid_payment = Payment.objects.create(
            deal=self.deal,
            amount=Decimal("500.00"),
            description="Paid",
            actual_date=timezone.now(),
        )
        self.authenticate(self.seller)
        response = self.api_client.delete(f"/api/v1/payments/{paid_payment.id}/")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)


class FinanceStatementTests(AuthenticatedAPITestCase):
    def setUp(self):
        super().setUp()
        self.seller = User.objects.create_user(  # pragma: allowlist secret
            username="seller", password="pass"  # pragma: allowlist secret
        )
        self.executor = User.objects.create_user(  # pragma: allowlist secret
            username="executor", password="pass"  # pragma: allowlist secret
        )
        client = Client.objects.create(name="Client")
        self.deal = Deal.objects.create(
            title="Statement Deal",
            client=client,
            seller=self.seller,
            executor=self.executor,
            status="open",
            stage_name="initial",
        )
        self.payment = Payment.objects.create(
            deal=self.deal, amount=Decimal("1000.00"), description="Initial"
        )
        self.income_record = FinancialRecord.objects.create(
            payment=self.payment, amount=Decimal("150.00"), description="Income"
        )
        self.zero_income_record = FinancialRecord.objects.create(
            payment=self.payment,
            amount=Decimal("0.00"),
            record_type=FinancialRecord.RecordType.INCOME,
            description="Zero income",
        )
        self.zero_expense_record = FinancialRecord.objects.create(
            payment=self.payment,
            amount=Decimal("0.00"),
            record_type=FinancialRecord.RecordType.EXPENSE,
            description="Zero expense",
        )
        self.expense_record = FinancialRecord.objects.create(
            payment=self.payment, amount=Decimal("-75.00"), description="Expense"
        )

    def test_create_statement_with_income_records(self):
        self.authenticate(self.seller)
        response = self.api_client.post(
            "/api/v1/finance_statements/",
            {
                "name": "Income Sheet",
                "statement_type": "income",
                "record_ids": [str(self.income_record.id)],
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.income_record.refresh_from_db()
        self.assertIsNotNone(self.income_record.statement_id)

    def test_cannot_add_expense_to_income_statement(self):
        self.authenticate(self.seller)
        response = self.api_client.post(
            "/api/v1/finance_statements/",
            {
                "name": "Income Sheet",
                "statement_type": "income",
                "record_ids": [str(self.expense_record.id)],
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        payload = response.json()
        errors = payload.get("record_ids", [])
        self.assertTrue(any("не подходит" in str(error) for error in errors))

    def test_can_add_zero_income_record_to_income_statement(self):
        self.authenticate(self.seller)
        response = self.api_client.post(
            "/api/v1/finance_statements/",
            {
                "name": "Income Sheet",
                "statement_type": "income",
                "record_ids": [str(self.zero_income_record.id)],
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.zero_income_record.refresh_from_db()
        self.assertIsNotNone(self.zero_income_record.statement_id)

    def test_can_add_zero_expense_record_to_expense_statement(self):
        self.authenticate(self.seller)
        response = self.api_client.post(
            "/api/v1/finance_statements/",
            {
                "name": "Expense Sheet",
                "statement_type": "expense",
                "record_ids": [str(self.zero_expense_record.id)],
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.zero_expense_record.refresh_from_db()
        self.assertIsNotNone(self.zero_expense_record.statement_id)

    def test_cannot_add_zero_income_record_to_expense_statement(self):
        self.authenticate(self.seller)
        response = self.api_client.post(
            "/api/v1/finance_statements/",
            {
                "name": "Expense Sheet",
                "statement_type": "expense",
                "record_ids": [str(self.zero_income_record.id)],
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        payload = response.json()
        errors = payload.get("record_ids", [])
        self.assertTrue(any("не подходит" in str(error) for error in errors))

    def test_record_cannot_be_used_in_multiple_statements(self):
        self.authenticate(self.seller)
        first = self.api_client.post(
            "/api/v1/finance_statements/",
            {
                "name": "Income Sheet",
                "statement_type": "income",
                "record_ids": [str(self.income_record.id)],
            },
            format="json",
        )
        self.assertEqual(first.status_code, status.HTTP_201_CREATED)
        response = self.api_client.post(
            "/api/v1/finance_statements/",
            {
                "name": "Another Sheet",
                "statement_type": "income",
                "record_ids": [str(self.income_record.id)],
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_cannot_create_statement_with_duplicate_name(self):
        self.authenticate(self.seller)
        response = self.api_client.post(
            "/api/v1/finance_statements/",
            {"name": "Duplicate Sheet", "statement_type": "income"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        duplicate = self.api_client.post(
            "/api/v1/finance_statements/",
            {"name": "Duplicate Sheet", "statement_type": "expense"},
            format="json",
        )

        self.assertEqual(duplicate.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("таким названием", str(duplicate.data.get("name")))

    def test_statement_duplicate_name_is_case_and_space_insensitive(self):
        self.authenticate(self.seller)
        response = self.api_client.post(
            "/api/v1/finance_statements/",
            {"name": "  Alice   Sheet  ", "statement_type": "income"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        duplicate = self.api_client.post(
            "/api/v1/finance_statements/",
            {"name": "alice sheet", "statement_type": "income"},
            format="json",
        )

        self.assertEqual(duplicate.status_code, status.HTTP_400_BAD_REQUEST)

    def test_can_update_statement_without_changing_name(self):
        self.authenticate(self.seller)
        statement = Statement.objects.create(
            name="Stable Sheet",
            statement_type="income",
            created_by=self.seller,
        )

        response = self.api_client.patch(
            f"/api/v1/finance_statements/{statement.id}/",
            {"comment": "Updated"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_cannot_rename_statement_to_duplicate_name(self):
        self.authenticate(self.seller)
        Statement.objects.create(
            name="Taken Sheet",
            statement_type="income",
            created_by=self.seller,
        )
        statement = Statement.objects.create(
            name="Editable Sheet",
            statement_type="expense",
            created_by=self.seller,
        )

        response = self.api_client.patch(
            f"/api/v1/finance_statements/{statement.id}/",
            {"name": "taken   sheet"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("таким названием", str(response.data.get("name")))

    def test_soft_deleted_statement_name_can_be_reused(self):
        self.authenticate(self.seller)
        statement = Statement.objects.create(
            name="Reusable Sheet",
            statement_type="income",
            created_by=self.seller,
        )
        statement.delete()

        response = self.api_client.post(
            "/api/v1/finance_statements/",
            {"name": "Reusable Sheet", "statement_type": "expense"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_cannot_edit_record_in_paid_statement(self):
        self.authenticate(self.seller)
        statement = Statement.objects.create(
            name="Paid Sheet",
            statement_type="income",
            paid_at=timezone.now().date(),
            created_by=self.seller,
        )
        self.income_record.statement = statement
        self.income_record.save(update_fields=["statement"])
        response = self.api_client.patch(
            f"/api/v1/financial_records/{self.income_record.id}/",
            {"amount": "120.00"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_cannot_delete_record_in_paid_statement(self):
        self.authenticate(self.seller)
        statement = Statement.objects.create(
            name="Paid Sheet",
            statement_type="income",
            paid_at=timezone.now().date(),
            created_by=self.seller,
        )
        self.income_record.statement = statement
        self.income_record.save(update_fields=["statement"])
        response = self.api_client.delete(
            f"/api/v1/financial_records/{self.income_record.id}/"
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_cannot_delete_record_in_draft_statement_requires_removal(self):
        self.authenticate(self.seller)
        statement = Statement.objects.create(
            name="Draft Sheet",
            statement_type="income",
            created_by=self.seller,
        )
        self.income_record.statement = statement
        self.income_record.save(update_fields=["statement"])
        response = self.api_client.delete(
            f"/api/v1/financial_records/{self.income_record.id}/"
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        message = str(response.data)
        self.assertIn("Сначала уберите её из состава ведомости", message)

    def test_delete_statement_unlinks_records(self):
        self.authenticate(self.seller)
        statement = Statement.objects.create(
            name="Draft Sheet",
            statement_type="income",
            created_by=self.seller,
        )
        self.income_record.statement = statement
        self.income_record.save(update_fields=["statement"])
        response = self.api_client.delete(f"/api/v1/finance_statements/{statement.id}/")
        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)
        self.income_record.refresh_from_db()
        self.assertIsNone(self.income_record.statement_id)

    def test_cannot_delete_paid_statement(self):
        self.authenticate(self.seller)
        statement = Statement.objects.create(
            name="Paid Sheet",
            statement_type="income",
            paid_at=timezone.now().date(),
            created_by=self.seller,
        )
        response = self.api_client.delete(f"/api/v1/finance_statements/{statement.id}/")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_cannot_mark_paid_without_date(self):
        self.authenticate(self.seller)
        statement = Statement.objects.create(
            name="Draft Sheet",
            statement_type="income",
            created_by=self.seller,
        )
        response = self.api_client.post(
            f"/api/v1/finance_statements/{statement.id}/mark-paid/",
            {},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_mark_paid_updates_record_dates(self):
        self.authenticate(self.seller)
        paid_at = timezone.now().date()
        statement = Statement.objects.create(
            name="To Pay",
            statement_type="income",
            created_by=self.seller,
        )
        self.income_record.statement = statement
        self.income_record.save(update_fields=["statement"])
        response = self.api_client.post(
            f"/api/v1/finance_statements/{statement.id}/mark-paid/",
            {"paid_at": paid_at.isoformat()},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        statement.refresh_from_db()
        self.income_record.refresh_from_db()
        self.assertEqual(statement.paid_at, paid_at)
        self.assertEqual(self.income_record.date, paid_at)

    def test_apply_amount_rub_updates_all_statement_records(self):
        self.authenticate(self.seller)
        statement = Statement.objects.create(
            name="Draft Sheet",
            statement_type="expense",
            created_by=self.seller,
        )
        self.expense_record.statement = statement
        self.zero_expense_record.statement = statement
        self.expense_record.save(update_fields=["statement"])
        self.zero_expense_record.save(update_fields=["statement"])

        response = self.api_client.post(
            f"/api/v1/finance_statements/{statement.id}/apply-amount/",
            {"mode": "rub", "value": "150"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        self.assertEqual(payload["updated"], 2)
        self.assertEqual(payload["unchanged"], 0)
        self.assertEqual(payload["skipped"], 0)
        self.expense_record.refresh_from_db()
        self.zero_expense_record.refresh_from_db()
        self.assertEqual(self.expense_record.amount, Decimal("-150.00"))
        self.assertEqual(self.zero_expense_record.amount, Decimal("-150.00"))

    def test_apply_amount_percent_uses_payment_paid_balance_and_skips_zero_balance(
        self,
    ):
        self.authenticate(self.seller)
        statement = Statement.objects.create(
            name="Percent Sheet",
            statement_type="income",
            created_by=self.seller,
        )
        second_payment = Payment.objects.create(
            deal=self.deal, amount=Decimal("2000.00"), description="Second"
        )
        first_record = FinancialRecord.objects.create(
            payment=self.payment, amount=Decimal("10.00"), statement=statement
        )
        second_record = FinancialRecord.objects.create(
            payment=second_payment, amount=Decimal("20.00"), statement=statement
        )
        zero_balance_record = FinancialRecord.objects.create(
            payment=Payment.objects.create(
                deal=self.deal, amount=Decimal("500.00"), description="No balance"
            ),
            amount=Decimal("30.00"),
            statement=statement,
        )
        FinancialRecord.objects.create(
            payment=self.payment,
            amount=Decimal("1000.00"),
            date=timezone.now().date(),
        )
        FinancialRecord.objects.create(
            payment=second_payment,
            amount=Decimal("2500.00"),
            date=timezone.now().date(),
        )

        response = self.api_client.post(
            f"/api/v1/finance_statements/{statement.id}/apply-amount/",
            {"mode": "percent", "value": "10"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        self.assertEqual(payload["updated"], 2)
        self.assertEqual(payload["unchanged"], 0)
        self.assertEqual(payload["skipped"], 1)
        self.assertEqual(payload["skipped_reasons"]["zero_balance"], 1)
        first_record.refresh_from_db()
        second_record.refresh_from_db()
        zero_balance_record.refresh_from_db()
        self.assertEqual(first_record.amount, Decimal("100.00"))
        self.assertEqual(second_record.amount, Decimal("250.00"))
        self.assertEqual(zero_balance_record.amount, Decimal("30.00"))

    def test_apply_amount_reports_unchanged_without_touching_record_updated_at(self):
        self.authenticate(self.seller)
        statement = Statement.objects.create(
            name="Unchanged Sheet",
            statement_type="income",
            created_by=self.seller,
        )
        self.income_record.statement = statement
        self.income_record.amount = Decimal("150.00")
        self.income_record.save(update_fields=["statement", "amount"])
        before_updated_at = self.income_record.updated_at

        response = self.api_client.post(
            f"/api/v1/finance_statements/{statement.id}/apply-amount/",
            {"mode": "rub", "value": "150"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        self.assertEqual(payload["updated"], 0)
        self.assertEqual(payload["unchanged"], 1)
        self.income_record.refresh_from_db()
        self.assertEqual(self.income_record.amount, Decimal("150.00"))
        self.assertEqual(self.income_record.updated_at, before_updated_at)

    def test_cannot_apply_amount_to_paid_statement(self):
        self.authenticate(self.seller)
        statement = Statement.objects.create(
            name="Paid Sheet",
            statement_type="income",
            paid_at=timezone.now().date(),
            created_by=self.seller,
        )
        self.income_record.statement = statement
        self.income_record.save(update_fields=["statement"])

        response = self.api_client.post(
            f"/api/v1/finance_statements/{statement.id}/apply-amount/",
            {"mode": "rub", "value": "150"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_executor_cannot_apply_amount_to_statement_records(self):
        self.authenticate(self.executor)
        statement = Statement.objects.create(
            name="Seller Sheet",
            statement_type="income",
            created_by=self.seller,
        )
        self.income_record.statement = statement
        self.income_record.save(update_fields=["statement"])

        response = self.api_client.post(
            f"/api/v1/finance_statements/{statement.id}/apply-amount/",
            {"mode": "rub", "value": "150"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_seller_sees_statement_listed_via_policy_deal_link(self):
        self.authenticate(self.seller)
        policy = Policy.objects.create(number="POLICY-LIST-SELLER", deal=self.deal)
        payment = Payment.objects.create(
            policy=policy,
            amount=Decimal("900.00"),
            description="Policy payment",
        )
        record = FinancialRecord.objects.create(
            payment=payment,
            amount=Decimal("180.00"),
            description="Policy income",
        )
        statement = Statement.objects.create(
            name="Policy-linked sheet",
            statement_type="income",
            created_by=self.seller,
        )
        record.statement = statement
        record.save(update_fields=["statement"])

        response = self.api_client.get("/api/v1/finance_statements/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = response.json().get("results", response.json())
        statement_ids = {item["id"] for item in results}
        self.assertIn(str(statement.id), statement_ids)

    def test_executor_sees_statement_listed_via_policy_deal_link(self):
        self.authenticate(self.executor)
        policy = Policy.objects.create(number="POLICY-LIST-EXECUTOR", deal=self.deal)
        payment = Payment.objects.create(
            policy=policy,
            amount=Decimal("950.00"),
            description="Executor policy payment",
        )
        record = FinancialRecord.objects.create(
            payment=payment,
            amount=Decimal("190.00"),
            description="Executor policy income",
        )
        statement = Statement.objects.create(
            name="Executor policy-linked sheet",
            statement_type="income",
            created_by=self.seller,
        )
        record.statement = statement
        record.save(update_fields=["statement"])

        response = self.api_client.get("/api/v1/finance_statements/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = response.json().get("results", response.json())
        statement_ids = {item["id"] for item in results}
        self.assertIn(str(statement.id), statement_ids)

    def test_paid_statement_via_policy_deal_is_returned_in_list(self):
        self.authenticate(self.seller)
        policy = Policy.objects.create(number="POLICY-LIST-PAID", deal=self.deal)
        payment = Payment.objects.create(
            policy=policy,
            amount=Decimal("1000.00"),
            description="Paid policy payment",
        )
        record = FinancialRecord.objects.create(
            payment=payment,
            amount=Decimal("200.00"),
            description="Paid policy income",
        )
        statement = Statement.objects.create(
            name="Paid policy-linked sheet",
            statement_type="income",
            created_by=self.seller,
            paid_at=timezone.now().date(),
        )
        record.statement = statement
        record.save(update_fields=["statement"])

        response = self.api_client.get("/api/v1/finance_statements/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = response.json().get("results", response.json())
        by_id = {item["id"]: item for item in results}
        self.assertIn(str(statement.id), by_id)
        self.assertEqual(
            by_id[str(statement.id)]["paid_at"], statement.paid_at.isoformat()
        )

    def test_statement_list_totals_ignore_soft_deleted_records(self):
        self.authenticate(self.seller)
        statement = Statement.objects.create(
            name="Statement totals",
            statement_type="income",
            created_by=self.seller,
        )
        active_income = FinancialRecord.objects.create(
            payment=self.payment,
            amount=Decimal("125.50"),
            description="Active income",
            statement=statement,
        )
        active_expense = FinancialRecord.objects.create(
            payment=self.payment,
            amount=Decimal("-25.50"),
            description="Active expense",
            statement=statement,
        )
        deleted_record = FinancialRecord.objects.create(
            payment=self.payment,
            amount=Decimal("900.00"),
            description="Deleted income",
            statement=statement,
        )
        deleted_record.delete()

        response = self.api_client.get("/api/v1/finance_statements/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = response.json().get("results", response.json())
        by_id = {item["id"]: item for item in results}
        self.assertIn(str(statement.id), by_id)
        self.assertEqual(by_id[str(statement.id)]["records_count"], 2)
        self.assertEqual(
            Decimal(str(by_id[str(statement.id)]["total_amount"])), Decimal("100.00")
        )
        active_income.refresh_from_db()
        active_expense.refresh_from_db()
        self.assertEqual(active_income.statement_id, statement.id)
        self.assertEqual(active_expense.statement_id, statement.id)

    def test_export_xlsx_creates_drive_file(self):
        self.authenticate(self.seller)
        fixed_now = timezone.make_aware(
            datetime(2026, 2, 8, 12, 34, 56), timezone=timezone.get_current_timezone()
        )

        uploaded = {}

        def fake_upload(folder_id, file_obj, file_name, mime_type):
            uploaded["folder_id"] = folder_id
            uploaded["file_name"] = file_name
            uploaded["mime_type"] = mime_type
            uploaded["bytes"] = file_obj.read()
            return {
                "id": "file123",
                "name": file_name,
                "mime_type": mime_type or "",
                "size": len(uploaded["bytes"]),
                "created_at": None,
                "modified_at": None,
                "web_view_link": "https://drive.test/file123",
                "is_folder": False,
            }

        with (
            patch(
                "apps.finances.signals.ensure_statement_folder",
                side_effect=lambda instance: instance.drive_folder_id,
            ),
            patch(
                "apps.finances.views.ensure_statement_folder", return_value="folder123"
            ),
            patch("apps.finances.views.upload_file_to_drive", side_effect=fake_upload),
            patch("apps.finances.views.timezone.now", return_value=fixed_now),
        ):
            statement = Statement.objects.create(
                name="Выплата Алиса",
                statement_type="income",
                created_by=self.seller,
                drive_folder_id="folder123",
            )
            self.payment.actual_date = fixed_now.date()
            self.payment.save(update_fields=["actual_date"])
            self.income_record.date = fixed_now.date()
            self.income_record.save(update_fields=["date"])
            self.expense_record.date = fixed_now.date()
            self.expense_record.save(update_fields=["date"])
            self.income_record.statement = statement
            self.income_record.save(update_fields=["statement"])
            response = self.api_client.post(
                f"/api/v1/finance_statements/{statement.id}/export-xlsx/",
                {},
                format="json",
            )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        self.assertEqual(payload["folder_id"], "folder123")
        self.assertEqual(payload["file"]["id"], "file123")

        self.assertTrue(
            uploaded["file_name"].startswith("Выплата Алиса_08_02_2026_12_34_56")
        )
        self.assertTrue(uploaded["file_name"].endswith(".xlsx"))

        workbook = load_workbook(filename=BytesIO(uploaded["bytes"]))
        ws = workbook.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertIn("Клиент / сделка", headers)
        self.assertIn("Доходы / расходы", headers)
        self.assertIn("Сумма, ₽", headers)
        header_index = {value: index + 1 for index, value in enumerate(headers)}

        operations_cell = ws.cell(row=2, column=header_index["Доходы / расходы"]).value
        self.assertCountEqual(
            operations_cell.splitlines(),
            [
                "Доход 150.00 ₽ · 08.02.2026",
                "Расход 75.00 ₽ · 08.02.2026",
            ],
        )
        self.assertEqual(
            ws.cell(row=2, column=header_index["Сальдо, ₽"]).value,
            75,
        )
        self.assertEqual(
            ws.cell(row=2, column=header_index["Сумма, ₽"]).value,
            150,
        )
        self.assertIn(
            "₽",
            ws.cell(row=2, column=header_index["Сальдо, ₽"]).number_format,
        )
        self.assertIn(
            "₽",
            ws.cell(row=2, column=header_index["Сумма, ₽"]).number_format,
        )
        self.assertNotIn(
            "Доход",
            str(ws.cell(row=2, column=header_index["Сальдо, ₽"]).value),
        )
        self.assertNotIn(
            "08.02.2026",
            str(ws.cell(row=2, column=header_index["Сумма, ₽"]).value),
        )
        self.assertNotIn(
            "+",
            str(ws.cell(row=2, column=header_index["Сумма, ₽"]).value),
        )
        self.assertNotIn(
            "-",
            str(ws.cell(row=2, column=header_index["Сумма, ₽"]).value),
        )
        workbook.close()

    def test_statement_drive_folder_is_created_even_when_name_exists(self):
        with patch(
            "apps.finances.signals.ensure_statement_folder",
            side_effect=lambda instance: instance.drive_folder_id,
        ):
            statement = Statement.objects.create(
                name="Drive Statement",
                statement_type="income",
                created_by=self.seller,
            )

        with (
            self.settings(GOOGLE_DRIVE_ROOT_FOLDER_ID="root-folder"),
            patch("apps.common.drive._ensure_folder", return_value="statements-root"),
            patch("apps.common.drive._find_folder", return_value={"id": "existing"}),
            patch(
                "apps.common.drive._make_folder", return_value="new-folder"
            ) as make_folder,
        ):
            folder_id = ensure_statement_folder(statement)

        self.assertEqual(folder_id, "new-folder")
        make_folder.assert_called_once_with("Drive Statement", "statements-root")
        statement.refresh_from_db()
        self.assertEqual(statement.drive_folder_id, "new-folder")


class FinancialRecordFilterTests(AuthenticatedAPITestCase):
    def setUp(self):
        super().setUp()
        self.seller = User.objects.create_user(  # pragma: allowlist secret
            username="seller", password="pass"  # pragma: allowlist secret
        )
        client = Client.objects.create(name="Search Client")
        self.deal = Deal.objects.create(
            title="Search Deal",
            client=client,
            seller=self.seller,
            status="open",
            stage_name="initial",
        )
        self.payment = Payment.objects.create(
            deal=self.deal, amount=Decimal("1000.00"), description="Payment seed"
        )
        self.policy = Policy.objects.create(number="FILTER-POLICY", deal=self.deal)
        self.policy_payment = Payment.objects.create(
            policy=self.policy,
            amount=Decimal("1200.00"),
            description="Policy payment seed",
        )
        self.income_record = FinancialRecord.objects.create(
            payment=self.payment, amount=Decimal("250.00"), note="AlphaNote"
        )
        self.expense_record = FinancialRecord.objects.create(
            payment=self.payment,
            amount=Decimal("-50.00"),
            description="Beta expense",
            date=timezone.now().date(),
        )
        self.policy_record = FinancialRecord.objects.create(
            payment=self.policy_payment,
            amount=Decimal("175.00"),
            note="Policy linked note",
        )
        self.statement = Statement.objects.create(
            name="Filter statement",
            statement_type="income",
            created_by=self.seller,
        )
        self.policy_record.statement = self.statement
        self.policy_record.save(update_fields=["statement"])

    def _extract_results(self, response):
        payload = response.json()
        return payload.get("results", payload)

    def test_filter_unpaid_only(self):
        self.authenticate(self.seller)
        response = self.api_client.get("/api/v1/financial_records/?unpaid_only=true")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        results = payload.get("results", payload)
        record_ids = {str(item["id"]) for item in results}
        self.assertIn(str(self.income_record.id), record_ids)
        self.assertNotIn(str(self.expense_record.id), record_ids)

    def test_filter_record_type_income(self):
        self.authenticate(self.seller)
        response = self.api_client.get("/api/v1/financial_records/?record_type=income")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        results = payload.get("results", payload)
        record_ids = {str(item["id"]) for item in results}
        self.assertIn(str(self.income_record.id), record_ids)
        self.assertNotIn(str(self.expense_record.id), record_ids)

    def test_filter_by_payment(self):
        self.authenticate(self.seller)
        response = self.api_client.get(
            "/api/v1/financial_records/",
            {"payment": str(self.policy_payment.id)},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        self.assertEqual([item["id"] for item in results], [str(self.policy_record.id)])

    def test_filter_by_policy(self):
        self.authenticate(self.seller)
        response = self.api_client.get(
            "/api/v1/financial_records/",
            {"policy": str(self.policy.id)},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        self.assertEqual([item["id"] for item in results], [str(self.policy_record.id)])

    def test_filter_by_sales_channel(self):
        target_channel = SalesChannel.objects.create(name="Target channel")
        other_channel = SalesChannel.objects.create(name="Other channel")
        other_policy = Policy.objects.create(number="OTHER-CHANNEL", deal=self.deal)
        self.policy.sales_channel = target_channel
        self.policy.save(update_fields=["sales_channel", "updated_at"])
        other_policy.sales_channel = other_channel
        other_policy.save(update_fields=["sales_channel", "updated_at"])
        other_payment = Payment.objects.create(
            policy=other_policy,
            amount=Decimal("900.00"),
            description="Other channel payment",
        )
        other_record = FinancialRecord.objects.create(
            payment=other_payment,
            amount=Decimal("90.00"),
            note="Other channel note",
        )

        self.authenticate(self.seller)
        response = self.api_client.get(
            "/api/v1/financial_records/",
            {"sales_channel": str(target_channel.id)},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        record_ids = {item["id"] for item in results}
        self.assertIn(str(self.policy_record.id), record_ids)
        self.assertNotIn(str(other_record.id), record_ids)
        self.assertNotIn(str(self.income_record.id), record_ids)

    def test_filter_by_payment_scheduled_date_range(self):
        early_payment = Payment.objects.create(
            policy=self.policy,
            amount=Decimal("900.00"),
            scheduled_date=date(2026, 3, 1),
        )
        target_payment = Payment.objects.create(
            policy=self.policy,
            amount=Decimal("950.00"),
            scheduled_date=date(2026, 3, 15),
        )
        late_payment = Payment.objects.create(
            policy=self.policy,
            amount=Decimal("990.00"),
            scheduled_date=date(2026, 4, 1),
        )
        early_record = FinancialRecord.objects.create(
            payment=early_payment, amount=Decimal("10.00")
        )
        target_record = FinancialRecord.objects.create(
            payment=target_payment, amount=Decimal("20.00")
        )
        late_record = FinancialRecord.objects.create(
            payment=late_payment, amount=Decimal("30.00")
        )

        self.authenticate(self.seller)
        response = self.api_client.get(
            "/api/v1/financial_records/",
            {
                "payment_scheduled_date_from": "2026-03-10",
                "payment_scheduled_date_to": "2026-03-20",
            },
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        record_ids = {item["id"] for item in results}
        self.assertIn(str(target_record.id), record_ids)
        self.assertNotIn(str(early_record.id), record_ids)
        self.assertNotIn(str(late_record.id), record_ids)

    def test_filter_by_deal_includes_policy_linked_records(self):
        self.authenticate(self.seller)
        response = self.api_client.get(
            "/api/v1/financial_records/",
            {"deal": str(self.deal.id)},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        returned_ids = {item["id"] for item in results}
        self.assertIn(str(self.income_record.id), returned_ids)
        self.assertIn(str(self.expense_record.id), returned_ids)
        self.assertIn(str(self.policy_record.id), returned_ids)

    def test_filter_by_statement(self):
        self.authenticate(self.seller)
        response = self.api_client.get(
            "/api/v1/financial_records/",
            {"statement": str(self.statement.id)},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        self.assertEqual([item["id"] for item in results], [str(self.policy_record.id)])

    def test_record_type_labels_are_readable(self):
        self.authenticate(self.seller)
        response = self.api_client.get("/api/v1/financial_records/")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        results = payload.get("results", payload)
        by_id = {str(item["id"]): item for item in results}
        self.assertEqual(by_id[str(self.income_record.id)]["record_type"], "Доход")
        self.assertEqual(by_id[str(self.expense_record.id)]["record_type"], "Расход")

    def test_record_type_is_source_of_truth_for_filters_and_labels(self):
        broken_expense = FinancialRecord.objects.create(
            payment=self.payment,
            amount=Decimal("-15.00"),
            note="Broken legacy expense",
        )
        FinancialRecord.objects.filter(id=broken_expense.id).update(
            amount=Decimal("15.00"),
            record_type=FinancialRecord.RecordType.EXPENSE,
        )

        self.authenticate(self.seller)
        response = self.api_client.get("/api/v1/financial_records/?record_type=expense")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        results = payload.get("results", payload)
        by_id = {str(item["id"]): item for item in results}

        self.assertIn(str(broken_expense.id), by_id)
        self.assertEqual(by_id[str(broken_expense.id)]["record_type"], "Расход")

    def test_search_applies_for_short_queries(self):
        self.authenticate(self.seller)
        response_short = self.api_client.get("/api/v1/financial_records/?search=Alph")
        self.assertEqual(response_short.status_code, status.HTTP_200_OK)
        payload_short = response_short.json()
        results_short = payload_short.get("results", payload_short)
        record_ids_short = {str(item["id"]) for item in results_short}
        self.assertIn(str(self.income_record.id), record_ids_short)
        self.assertNotIn(str(self.expense_record.id), record_ids_short)

        response = self.api_client.get("/api/v1/financial_records/?search=AlphaNote")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        results = payload.get("results", payload)
        record_ids = {str(item["id"]) for item in results}
        self.assertIn(str(self.income_record.id), record_ids)
        self.assertNotIn(str(self.expense_record.id), record_ids)

    def test_search_keeps_all_supported_fields(self):
        insured_client = Client.objects.create(name="InsuredSearchToken")
        insurance_type = InsuranceType.objects.create(name="TypeSearchToken")
        sales_channel = SalesChannel.objects.create(name="ChannelSearchToken")
        self.policy.client = self.deal.client
        self.policy.insured_client = insured_client
        self.policy.insurance_type = insurance_type
        self.policy.sales_channel = sales_channel
        self.policy.save(
            update_fields=[
                "client",
                "insured_client",
                "insurance_type",
                "sales_channel",
                "updated_at",
            ]
        )
        self.policy_record.description = "RecordDescriptionToken"
        self.policy_record.source = "RecordSourceToken"
        self.policy_record.note = "RecordNoteToken"
        self.policy_record.save(update_fields=["description", "source", "note"])

        search_terms = (
            "FILTER-POLICY",
            "Search Client",
            "InsuredSearchToken",
            "TypeSearchToken",
            "ChannelSearchToken",
            "Search Deal",
            "Policy payment seed",
            "RecordDescriptionToken",
            "RecordSourceToken",
            "RecordNoteToken",
        )

        self.authenticate(self.seller)
        for search_term in search_terms:
            with self.subTest(search_term=search_term):
                response = self.api_client.get(
                    "/api/v1/financial_records/", {"search": search_term}
                )
                self.assertEqual(response.status_code, status.HTTP_200_OK)
                returned_ids = {item["id"] for item in self._extract_results(response)}
                self.assertIn(str(self.policy_record.id), returned_ids)

    def test_financial_record_list_query_count_does_not_grow_per_policy(self):
        one_client = Client.objects.create(name="Query One Client")
        one_policy = Policy.objects.create(
            number="QUERY-ONE",
            deal=self.deal,
            client=one_client,
            insured_client=one_client,
        )
        one_payment = Payment.objects.create(
            policy=one_policy,
            amount=Decimal("10.00"),
            description="Query one payment",
        )
        FinancialRecord.objects.create(payment=one_payment, amount=Decimal("1.00"))

        for index in range(20):
            client = Client.objects.create(name=f"Query Many Client {index}")
            policy = Policy.objects.create(
                number=f"QUERY-MANY-{index}",
                deal=self.deal,
                client=client,
                insured_client=client,
            )
            payment = Payment.objects.create(
                policy=policy,
                amount=Decimal("10.00"),
                description=f"Query many payment {index}",
            )
            FinancialRecord.objects.create(payment=payment, amount=Decimal("1.00"))

        self.authenticate(self.seller)
        with CaptureQueriesContext(connection) as one_context:
            one_response = self.api_client.get(
                "/api/v1/financial_records/", {"search": "QUERY-ONE"}
            )
        with CaptureQueriesContext(connection) as many_context:
            many_response = self.api_client.get(
                "/api/v1/financial_records/", {"search": "QUERY-MANY"}
            )

        self.assertEqual(one_response.status_code, status.HTTP_200_OK)
        self.assertEqual(many_response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(self._extract_results(one_response)), 1)
        self.assertEqual(len(self._extract_results(many_response)), 20)
        self.assertEqual(len(one_context), len(many_context))

    def test_filter_paid_balance_not_zero(self):
        self.authenticate(self.seller)
        paid_income = FinancialRecord.objects.create(
            payment=self.payment, amount=Decimal("60.00"), date=timezone.now().date()
        )
        response = self.api_client.get(
            "/api/v1/financial_records/?paid_balance_not_zero=true"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        results = payload.get("results", payload)
        record_ids = {str(item["id"]) for item in results}
        self.assertIn(str(paid_income.id), record_ids)

    def test_ordering_by_payment_paid_balance_works(self):
        self.authenticate(self.seller)
        payment_low = self.payment
        payment_high = Payment.objects.create(
            deal=self.deal, amount=Decimal("2000.00"), description="Other payment"
        )

        # payment_low already has a paid expense record in setUp -> contributes to balance.
        low_paid = FinancialRecord.objects.create(
            payment=payment_low, amount=Decimal("10.00"), date=timezone.now().date()
        )
        high_paid = FinancialRecord.objects.create(
            payment=payment_high, amount=Decimal("30.00"), date=timezone.now().date()
        )

        response = self.api_client.get(
            "/api/v1/financial_records/?ordering=-payment_paid_balance"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        results = payload.get("results", payload)
        returned_ids = [str(item["id"]) for item in results]

        # Records for payment_high should come before records for payment_low.
        self.assertLess(
            returned_ids.index(str(high_paid.id)), returned_ids.index(str(low_paid.id))
        )

    def test_ordering_by_amount_works(self):
        self.authenticate(self.seller)
        low = FinancialRecord.objects.create(
            payment=self.payment, amount=Decimal("10.00")
        )
        high = FinancialRecord.objects.create(
            payment=self.payment, amount=Decimal("20.00")
        )

        response = self.api_client.get("/api/v1/financial_records/?ordering=amount")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        results = payload.get("results", payload)
        returned_ids = [str(item["id"]) for item in results]
        self.assertLess(
            returned_ids.index(str(low.id)), returned_ids.index(str(high.id))
        )

    def test_ordering_by_record_comment_sort_works(self):
        self.authenticate(self.seller)
        a = FinancialRecord.objects.create(
            payment=self.payment, amount=Decimal("10.00"), note="AAA"
        )
        b = FinancialRecord.objects.create(
            payment=self.payment, amount=Decimal("20.00"), note="BBB"
        )

        response = self.api_client.get(
            "/api/v1/financial_records/?ordering=record_comment_sort"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        results = payload.get("results", payload)
        returned_ids = [str(item["id"]) for item in results]
        self.assertLess(returned_ids.index(str(a.id)), returned_ids.index(str(b.id)))

    def test_ordering_by_payment_is_paid_works(self):
        self.authenticate(self.seller)
        paid_payment = Payment.objects.create(
            deal=self.deal,
            amount=Decimal("1500.00"),
            description="Paid payment",
            actual_date=timezone.now().date(),
        )
        unpaid_record = FinancialRecord.objects.create(
            payment=self.payment, amount=Decimal("10.00")
        )
        paid_record = FinancialRecord.objects.create(
            payment=paid_payment, amount=Decimal("10.00")
        )

        response = self.api_client.get(
            "/api/v1/financial_records/?ordering=-payment_is_paid"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        results = payload.get("results", payload)
        returned_ids = [str(item["id"]) for item in results]
        self.assertLess(
            returned_ids.index(str(paid_record.id)),
            returned_ids.index(str(unpaid_record.id)),
        )

    def test_ordering_by_payment_scheduled_date_ignores_actual_date(self):
        older_scheduled_payment = Payment.objects.create(
            policy=self.policy,
            amount=Decimal("1000.00"),
            scheduled_date=date(2026, 3, 1),
            actual_date=date(2026, 5, 1),
        )
        newer_scheduled_payment = Payment.objects.create(
            policy=self.policy,
            amount=Decimal("1000.00"),
            scheduled_date=date(2026, 4, 1),
            actual_date=date(2026, 1, 1),
        )
        no_date_payment = Payment.objects.create(
            policy=self.policy,
            amount=Decimal("1000.00"),
        )
        older_record = FinancialRecord.objects.create(
            payment=older_scheduled_payment, amount=Decimal("10.00")
        )
        newer_record = FinancialRecord.objects.create(
            payment=newer_scheduled_payment, amount=Decimal("20.00")
        )
        no_date_record = FinancialRecord.objects.create(
            payment=no_date_payment, amount=Decimal("30.00")
        )

        self.authenticate(self.seller)
        response = self.api_client.get(
            "/api/v1/financial_records/",
            {
                "payment_scheduled_date_from": "2026-01-01",
                "ordering": "payment_scheduled_date_is_null,payment_scheduled_date",
            },
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        returned_ids = [item["id"] for item in self._extract_results(response)]
        self.assertLess(
            returned_ids.index(str(older_record.id)),
            returned_ids.index(str(newer_record.id)),
        )

        response_desc = self.api_client.get(
            "/api/v1/financial_records/",
            {
                "ordering": (
                    "payment_scheduled_date_is_null,"
                    "-payment_scheduled_date,-created_at"
                ),
            },
        )
        self.assertEqual(response_desc.status_code, status.HTTP_200_OK)
        returned_ids_desc = [
            item["id"] for item in self._extract_results(response_desc)
        ]
        self.assertLess(
            returned_ids_desc.index(str(newer_record.id)),
            returned_ids_desc.index(str(older_record.id)),
        )
        self.assertLess(
            returned_ids_desc.index(str(older_record.id)),
            returned_ids_desc.index(str(no_date_record.id)),
        )

    def test_export_xlsx_uses_filters_and_includes_payment_scheduled_date(self):
        channel = SalesChannel.objects.create(name="Export channel")
        self.policy.sales_channel = channel
        self.policy.save(update_fields=["sales_channel", "updated_at"])
        self.policy_payment.scheduled_date = date(2026, 3, 15)
        self.policy_payment.save(update_fields=["scheduled_date", "updated_at"])

        self.authenticate(self.seller)
        response = self.api_client.get(
            "/api/v1/financial_records/export-xlsx/",
            {"sales_channel": str(channel.id)},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        self.assertIn("Дата платежа", headers)
        scheduled_date_column = headers.index("Дата платежа") + 1
        policy_number_column = headers.index("Номер полиса") + 1

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        policy_numbers = {row[policy_number_column - 1] for row in rows}
        scheduled_dates = {row[scheduled_date_column - 1] for row in rows}
        self.assertIn("FILTER-POLICY", policy_numbers)
        self.assertNotIn("-", policy_numbers)
        self.assertIn("15.03.2026", scheduled_dates)


class FinancialRecordPaidBalanceTests(AuthenticatedAPITestCase):
    def setUp(self):
        super().setUp()
        self.seller = User.objects.create_user(  # pragma: allowlist secret
            username="seller", password="pass"  # pragma: allowlist secret
        )
        self.viewer = User.objects.create_user(  # pragma: allowlist secret
            username="viewer", password="pass"  # pragma: allowlist secret
        )
        client = Client.objects.create(name="Saldo Client")
        self.deal = Deal.objects.create(
            title="Saldo Deal",
            client=client,
            seller=self.seller,
            status="open",
            stage_name="initial",
        )
        self.policy = Policy.objects.create(number="SALDO-POLICY", deal=self.deal)

    def _extract_results(self, response):
        payload = response.json()
        return payload.get("results", payload)

    def test_paid_balance_is_not_multiplied_by_visibility_joins(self):
        self.deal.visible_users.add(self.viewer)
        Task.objects.create(
            title="Visibility task 1",
            deal=self.deal,
            assignee=self.seller,
        )
        Task.objects.create(
            title="Visibility task 2",
            deal=self.deal,
            assignee=self.seller,
        )
        payment = Payment.objects.create(
            deal=self.deal,
            policy=self.policy,
            amount=Decimal("200032.00"),
            description="Tagiev payment",
            actual_date=timezone.now().date(),
        )
        paid_expense = FinancialRecord.objects.create(
            payment=payment,
            amount=Decimal("-24000.00"),
            date=timezone.now().date(),
            note="Скидка",
        )
        unpaid_expense = FinancialRecord.objects.create(
            payment=payment,
            amount=Decimal("-1.00"),
            note="Расход исполнителю",
        )
        income = FinancialRecord.objects.create(
            payment=payment,
            amount=Decimal("1.00"),
            note="Комиссия",
        )

        self.authenticate(self.seller)
        response = self.api_client.get("/api/v1/financial_records/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        by_id = {item["id"]: item for item in results}
        expected_balance = "-24000.00"
        for record in (paid_expense, unpaid_expense, income):
            self.assertEqual(
                by_id[str(record.id)]["payment_paid_balance"], expected_balance
            )

    def test_paid_balance_stays_correct_without_visibility_joins(self):
        payment = Payment.objects.create(
            deal=self.deal,
            policy=self.policy,
            amount=Decimal("5000.00"),
            description="Simple payment",
            actual_date=timezone.now().date(),
        )
        paid_income = FinancialRecord.objects.create(
            payment=payment,
            amount=Decimal("150.00"),
            date=timezone.now().date(),
        )
        unpaid_income = FinancialRecord.objects.create(
            payment=payment,
            amount=Decimal("50.00"),
        )

        self.authenticate(self.seller)
        response = self.api_client.get("/api/v1/financial_records/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = self._extract_results(response)
        by_id = {item["id"]: item for item in results}
        self.assertEqual(by_id[str(paid_income.id)]["payment_paid_balance"], "150.00")
        self.assertEqual(by_id[str(unpaid_income.id)]["payment_paid_balance"], "150.00")


class FinanceStatementRemoveRecordsTests(AuthenticatedAPITestCase):
    def setUp(self):
        super().setUp()
        self.seller = User.objects.create_user(  # pragma: allowlist secret
            username="seller", password="pass"  # pragma: allowlist secret
        )
        client = Client.objects.create(name="Remove Client")
        self.deal = Deal.objects.create(
            title="Remove Deal",
            client=client,
            seller=self.seller,
            status="open",
            stage_name="initial",
        )
        self.payment = Payment.objects.create(
            deal=self.deal, amount=Decimal("1000.00"), description="Seed"
        )
        self.income_record = FinancialRecord.objects.create(
            payment=self.payment, amount=Decimal("120.00")
        )
        self.statement = Statement.objects.create(
            name="Remove Sheet",
            statement_type="income",
            created_by=self.seller,
        )
        self.income_record.statement = self.statement
        self.income_record.save(update_fields=["statement"])

    def test_can_remove_record_from_statement(self):
        self.authenticate(self.seller)
        response = self.api_client.post(
            f"/api/v1/finance_statements/{self.statement.id}/remove-records/",
            {"record_ids": [str(self.income_record.id)]},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.income_record.refresh_from_db()
        self.assertIsNone(self.income_record.statement_id)


class StatementDriveDownloadTests(AuthenticatedAPITestCase):
    def setUp(self):
        super().setUp()
        self.seller = User.objects.create_user(  # pragma: allowlist secret
            username="statement-drive-seller",
            password="pass",  # pragma: allowlist secret
        )
        self.statement = Statement.objects.create(
            name="Drive Statement",
            statement_type="income",
            created_by=self.seller,
            drive_folder_id="statement-folder",
        )
        self.authenticate(self.seller)

    def test_downloads_nested_file_from_statement_subfolder(self):
        file_map = {
            "nested-file": {
                "id": "nested-file",
                "name": "nested.pdf",
                "mime_type": "application/pdf",
                "is_folder": False,
                "parent_id": "subfolder-1",
            }
        }

        with (
            patch(
                "apps.finances.views.build_drive_file_tree_map",
                return_value=file_map,
            ) as tree_mock,
            patch(
                "apps.finances.views.download_drive_file",
                return_value=b"nested-pdf",
            ) as download_mock,
        ):
            response = self.api_client.post(
                f"/api/v1/finance_statements/{self.statement.id}/drive-files/download/",
                {"file_ids": ["nested-file"]},
                format="json",
            )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.content, b"nested-pdf")
        self.assertIn("nested.pdf", response["Content-Disposition"])
        tree_mock.assert_called_once_with("statement-folder")
        download_mock.assert_called_once_with("nested-file")

    def test_downloads_zip_with_files_from_nested_statement_folders(self):
        file_map = {
            "nested-file": {
                "id": "nested-file",
                "name": "nested.pdf",
                "mime_type": "application/pdf",
                "is_folder": False,
                "parent_id": "subfolder-1",
            },
            "root-file": {
                "id": "root-file",
                "name": "root.pdf",
                "mime_type": "application/pdf",
                "is_folder": False,
                "parent_id": "statement-folder",
            },
        }

        with (
            patch(
                "apps.finances.views.build_drive_file_tree_map",
                return_value=file_map,
            ),
            patch(
                "apps.finances.views.download_drive_file",
                side_effect=[b"nested-pdf", b"root-pdf"],
            ),
        ):
            response = self.api_client.post(
                f"/api/v1/finance_statements/{self.statement.id}/drive-files/download/",
                {"file_ids": ["nested-file", "root-file"]},
                format="json",
            )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        archive = zipfile.ZipFile(BytesIO(response.content))
        self.assertEqual(sorted(archive.namelist()), ["nested.pdf", "root.pdf"])
        self.assertEqual(archive.read("nested.pdf"), b"nested-pdf")
        self.assertEqual(archive.read("root.pdf"), b"root-pdf")

    def test_rejects_folder_selection_found_in_nested_statement_tree(self):
        file_map = {
            "nested-folder": {
                "id": "nested-folder",
                "name": "folder",
                "mime_type": "application/vnd.google-apps.folder",
                "is_folder": True,
                "parent_id": "statement-folder",
            }
        }

        with patch(
            "apps.finances.views.build_drive_file_tree_map",
            return_value=file_map,
        ) as tree_mock:
            response = self.api_client.post(
                f"/api/v1/finance_statements/{self.statement.id}/drive-files/download/",
                {"file_ids": ["nested-folder"]},
                format="json",
            )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            response.data["detail"],
            "Скачивание папок не поддерживается. Выберите только файлы.",
        )
        tree_mock.assert_called_once_with("statement-folder")
