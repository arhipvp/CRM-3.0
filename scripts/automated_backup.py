#!/usr/bin/env python3
"""
Automated CRM 3.0 backup.

Runs every three hours (cron / Task Scheduler) and produces:
  * full pg_dump
  * Excel workbook with one sheet per table
  * incremental copies of backend/media
  * tiered retention buckets: 3 hours, 1 day, 1 week
"""

from __future__ import annotations

import logging
import os
import shutil
import subprocess
import sys
import uuid
from datetime import datetime, date, time, timedelta
from pathlib import Path
from typing import Iterable

try:
    import psycopg
    from psycopg.rows import dict_row
    from psycopg.sql import Identifier, SQL
except ImportError as exc:  # pragma: no cover
    raise RuntimeError("Install psycopg[binary] before running this script.") from exc

try:
    from openpyxl import Workbook
except ImportError as exc:  # pragma: no cover
    raise RuntimeError("Install openpyxl (see backend/requirements.txt).") from exc

ROOT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT_DIR / "backend"
BACKUP_ROOT = ROOT_DIR / "backups"
PROJECT_FILES_SRC = BACKEND_DIR / "media"
PROJECT_FILES_DST = BACKUP_ROOT / "project_files"

HOURLY_RETENTION_HOURS = 3
DAILY_RETENTION_HOURS = 24
WEEKLY_RETENTION_HOURS = 7 * 24

DAILY_PROMOTE_HOURS = 24
WEEKLY_PROMOTE_HOURS = 7 * 24

LOGGER = logging.getLogger("crm_backup")


def setup_django_settings():
    """Add backend to PATH, chdir there, and configure Django."""
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
    sys.path.insert(0, str(BACKEND_DIR))
    os.chdir(BACKEND_DIR)
    import django  # noqa: WPS433

    django.setup()
    from django.conf import settings

    return settings


def build_db_params(settings):
    """Build psycopg-friendly connection arguments based on Django settings."""
    config = settings.DATABASES["default"]
    params = {"dbname": config["NAME"]}
    for source_key, target_key in (
        ("USER", "user"),
        ("PASSWORD", "password"),
        ("HOST", "host"),
        ("PORT", "port"),
    ):
        value = config.get(source_key)
        if value:
            params[target_key] = value
    # psycopg expects str or None for port
    if "port" in params:
        try:
            params["port"] = int(params["port"])
        except ValueError:
            params["port"] = params["port"]
    return params


def ensure_dirs(*paths: Path):
    for path in paths:
        path.mkdir(parents=True, exist_ok=True)


def run_pg_dump(settings, output_path: Path):
    """Dump the database with pg_dump or fallback to docker compose."""
    db_conf = settings.DATABASES["default"]
    pg_dump_cmd = shutil.which("pg_dump")
    if pg_dump_cmd:
        LOGGER.info("pg_dump found locally, saving SQL to %s", output_path)
        args = [pg_dump_cmd]
        if db_conf.get("HOST"):
            args.extend(["-h", db_conf["HOST"]])
        if db_conf.get("PORT"):
            args.extend(["-p", str(db_conf["PORT"])])
        if db_conf.get("USER"):
            args.extend(["-U", db_conf["USER"]])
        args.append(db_conf["NAME"])
        env = os.environ.copy()
        if db_conf.get("PASSWORD"):
            env["PGPASSWORD"] = db_conf["PASSWORD"]
        with output_path.open("wb") as stream:
            subprocess.run(args, check=True, stdout=stream, env=env)
        return

    LOGGER.info("pg_dump missing, trying docker compose")
    docker_cmd = detect_docker_compose()
    cmd = docker_cmd + [
        "exec",
        "-T",
        "db",
        "pg_dump",
        "-U",
        db_conf.get("USER") or "crm3",
        db_conf["NAME"],
    ]
    with output_path.open("wb") as stream:
        subprocess.run(cmd, check=True, stdout=stream, cwd=ROOT_DIR)


def detect_docker_compose() -> list[str]:
    checks = [
        (["docker", "compose"], ["docker", "compose", "version"]),
        (["docker-compose"], ["docker-compose", "version"]),
    ]
    for command, version_cmd in checks:
        try:
            subprocess.run(
                version_cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
            )
            return command
        except (FileNotFoundError, subprocess.CalledProcessError):
            continue
    raise RuntimeError("Docker compose is not available (docker compose / docker-compose).")


def export_database_to_excel(db_params, output_path: Path):
    """Export each public table into a separate Excel worksheet."""
    LOGGER.info("Collecting data for Excel (%s)", output_path.name)
    workbook = Workbook()
    worksheet_names: set[str] = set()

    with psycopg.connect(**db_params) as conn:
        with conn.cursor(row_factory=dict_row) as cursor:
            cursor.execute(
                """
                SELECT table_name
                FROM information_schema.tables
                WHERE table_schema = 'public'
                  AND table_type = 'BASE TABLE'
                ORDER BY table_name
                """
            )
            tables = [row["table_name"] for row in cursor.fetchall()]
            for table in tables:
                cursor.execute(SQL("SELECT * FROM {}").format(Identifier(table)))
                columns = getattr(cursor, "column_names", None)
                if columns is None:
                    columns = [col.name for col in cursor.description] if cursor.description else []
                sheet_title = sanitize_sheet_name(table, worksheet_names)
                worksheet = workbook.create_sheet(title=sheet_title)
                worksheet.append(columns or ["(no columns)"])
                for record in cursor:
                    worksheet.append([normalize_value(record[col]) for col in columns])

    if workbook.sheetnames == ["Sheet"]:
        workbook.active.title = "data"
    workbook.save(output_path)


def sanitize_sheet_name(source: str, used: set[str]) -> str:
    base = source[:31]
    candidate = base or "table"
    counter = 1
    while candidate in used:
        suffix = f"_{counter}"
        limit = 31 - len(suffix)
        candidate = (base[:limit] if len(base) > limit else base) + suffix
        counter += 1
    used.add(candidate)
    return candidate


def normalize_value(value):
    if value is None:
        return value
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, time):
        return value.replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, date):
        return value
    if isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, uuid.UUID):
        return str(value)
    if isinstance(value, bytes):
        return value.decode("utf-8", "ignore")
    return str(value)


def copy_new_project_files(source: Path, destination: Path):
    if not source.exists():
        LOGGER.warning("Project files directory not found: %s", source)
        return
    ensure_dirs(destination)
    copied = 0
    for entry in source.rglob("*"):
        if not entry.is_file():
            continue
        relative = entry.relative_to(source)
        target_path = destination / relative
        if target_path.exists():
            continue
        target_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(entry, target_path)
        copied += 1
    LOGGER.info("New user files copied: %d", copied)


def prune_directory(directory: Path, hours: int, now: datetime):
    if not directory.exists():
        return
    limit = now - timedelta(hours=hours)
    for file in directory.iterdir():
        if not file.is_file():
            continue
        if datetime.fromtimestamp(file.stat().st_mtime) < limit:
            LOGGER.debug("Removing stale file %s", file.name)
            file.unlink()


def latest_modification(directory: Path) -> datetime | None:
    if not directory.exists():
        return None
    times = [
        datetime.fromtimestamp(entry.stat().st_mtime)
        for entry in directory.iterdir()
        if entry.is_file()
    ]
    return max(times) if times else None


def promote_snapshot(sources: Iterable[Path], tier_dir: Path, interval_hours: int, now: datetime):
    ensure_dirs(tier_dir)
    modified = latest_modification(tier_dir)
    if modified and (now - modified) < timedelta(hours=interval_hours):
        return
    for src in sources:
        destination = tier_dir / src.name
        shutil.copy2(src, destination)
        LOGGER.info("Snapshot %s -> %s", src.name, destination.relative_to(ROOT_DIR))


def main():
    logging.basicConfig(level=logging.INFO, format="[%(asctime)s] %(message)s")
    settings = setup_django_settings()
    timestamp = datetime.now()
    snapshot_key = timestamp.strftime("%Y%m%d_%H%M%S")
    hourly_dir = BACKUP_ROOT / "hourly"
    ensure_dirs(hourly_dir, PROJECT_FILES_DST)

    sql_path = hourly_dir / f"db_dump_{snapshot_key}.sql"
    excel_path = hourly_dir / f"db_dump_{snapshot_key}.xlsx"

    run_pg_dump(settings, sql_path)
    export_database_to_excel(build_db_params(settings), excel_path)
    copy_new_project_files(PROJECT_FILES_SRC, PROJECT_FILES_DST)

    promote_snapshot(
        [sql_path, excel_path],
        BACKUP_ROOT / "daily",
        DAILY_PROMOTE_HOURS,
        timestamp,
    )
    promote_snapshot(
        [sql_path, excel_path],
        BACKUP_ROOT / "weekly",
        WEEKLY_PROMOTE_HOURS,
        timestamp,
    )

    prune_directory(hourly_dir, HOURLY_RETENTION_HOURS, timestamp)
    prune_directory(BACKUP_ROOT / "daily", DAILY_RETENTION_HOURS, timestamp)
    prune_directory(BACKUP_ROOT / "weekly", WEEKLY_RETENTION_HOURS, timestamp)

    LOGGER.info("Backup complete: hourly=%s, Excel=%s", sql_path.name, excel_path.name)


if __name__ == "__main__":
    main()
