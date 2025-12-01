from __future__ import annotations

import argparse
import json
import os
import sys
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping

import django
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

BASE_DIR = Path(__file__).resolve().parents[1]
if (BASE_DIR / "backend").exists():
    BACKEND_DIR = BASE_DIR / "backend"
else:
    BACKEND_DIR = BASE_DIR


def _load_backend_env() -> None:
    env_path = BACKEND_DIR / ".env"
    if not env_path.exists():
        return

    with env_path.open(encoding="utf-8") as stream:
        for line in stream:
            raw = line.strip()
            if not raw or raw.startswith("#") or "=" not in raw:
                continue
            key, value = raw.split("=", 1)
            key = key.strip().lstrip("\ufeff")
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


sys.path.insert(0, str(BACKEND_DIR))
os.chdir(BACKEND_DIR)

_load_backend_env()

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

from django.apps import apps
from django.core.exceptions import FieldDoesNotExist
from django.db import DataError, connection, models


@dataclass(frozen=True)
class SheetSpec:
    model_path: str
    field_map: Mapping[str, str]
    defaults: Mapping[str, Any] = field(default_factory=dict)
    post_process: Callable[[dict[str, Any], Mapping[str, Any]], None] | None = None


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Импортирует CRM-данные из Excel-шаблона (смотрите scripts/templates/business_data_template_new.xlsx)."
    )
    parser.add_argument("path", help="Путь к Excel-файлу (.xlsx)")
    parser.add_argument(
        "--sheet", help="Импортировать только указанный лист (по умолчанию все)"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Прогон импорта без сохранения строк в базу",
    )
    parser.add_argument(
        "--clear",
        action="store_true",
        help="Очистить связанные таблицы перед импортом",
    )
    return parser.parse_args()


def _normalize_sheet_name(name: str) -> str:
    return name.strip().lower()


def _iter_rows(sheet: Worksheet) -> Iterable[tuple[int, dict[str, Any]]]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(cell).strip().lower() if cell else "" for cell in header_row]

    for row_index, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if all(cell in (None, "", False) for cell in row):
            continue

        payload: dict[str, Any] = {}
        for header, value in zip(headers, row):
            if not header:
                continue
            payload[header] = value

        if payload:
            yield row_index, payload


def _coerce_value(field: models.Field, value: Any, treat_as_id: bool) -> Any:
    if value is None:
        if isinstance(field, (models.CharField, models.TextField)) and not field.null:
            value = ""
        else:
            return None
    if isinstance(value, bool):
        return None
    if isinstance(value, str) and not value.strip():
        if isinstance(field, (models.CharField, models.TextField)) and not field.null:
            return ""
        return None

    if isinstance(field, models.DecimalField):
        return Decimal(str(value))

    if isinstance(field, models.DateField) and not isinstance(
        field, models.DateTimeField
    ):
        return _parse_date(value)

    if isinstance(field, models.DateTimeField):
        return _parse_datetime(value)

    if isinstance(field, models.ForeignKey):
        return _resolve_fk_value(field, value)

    if treat_as_id and isinstance(value, (int, float)):
        return int(value)

    if isinstance(field, models.BooleanField):
        if isinstance(value, str):
            lowered = value.strip().lower()
            if lowered in {"true", "1", "yes"}:
                return True
            if lowered in {"false", "0", "no"}:
                return False
        return bool(value)

    result = field.to_python(value)
    if (
        isinstance(field, models.CharField)
        and field.max_length
        and isinstance(result, str)
    ):
        if len(result) > field.max_length:
            print(
                "trimming",
                field.name,
                "from",
                len(result),
                "to",
                field.max_length,
                "characters",
            )
            result = result[: field.max_length]
    if (
        result is None
        and isinstance(field, (models.CharField, models.TextField))
        and not field.null
    ):
        return ""
    return result


CLEAR_MODEL_ORDER = [
    "tasks.Task",
    "finances.FinancialRecord",
    "finances.Payment",
    "policies.Policy",
    "documents.Document",
    "notes.Note",
    "chat.ChatMessage",
    "deals.Quote",
    "deals.Deal",
    "clients.Client",
]


def _clear_tables() -> None:
    """Удаляет данные из таблиц в порядке, соответствующем зависимостям моделей."""
    with connection.cursor() as cursor:
        for model_path in CLEAR_MODEL_ORDER:
            model = apps.get_model(model_path)
            table = connection.ops.quote_name(model._meta.db_table)
            cursor.execute(f"DELETE FROM {table};")
    CREATED_POLICY_IDS.clear()
    CREATED_PAYMENT_IDS.clear()
    CREATED_DEAL_IDS.clear()


def _resolve_fk_value(field: models.ForeignKey, value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, str) and not value.strip():
        return None

    related = field.remote_field.model
    if related._meta.label_lower == "clients.client":
        mapped = _map_to_client_uuid(value)
        if mapped:
            return mapped

    if related._meta.label_lower == "deals.deal":
        mapped = _map_to_deal_uuid(value)
        if mapped:
            return mapped

    if related._meta.label_lower == "policies.policy":
        mapped = _map_to_policy_uuid(value)
        if mapped:
            return mapped

    if related._meta.label_lower == "finances.payment":
        mapped = _map_to_payment_uuid(value)
        if mapped:
            return mapped

    if isinstance(value, (int, float)):
        return int(value)

    text = str(value).strip()
    if text.isdigit():
        return int(text)

    for attr in ("name", "title", "number"):
        if hasattr(related, attr):
            lookup = {attr: text}
            result = related.objects.filter(**lookup).first()
            if result:
                return result.pk
    created_pk = _auto_create_related(related, text)
    if created_pk:
        return created_pk

    raise ValueError(f"Не найдено {related._meta.verbose_name} '{value}'")


def _auto_create_related(related: type[models.Model], text: str) -> int | None:
    key = related._meta.label_lower
    config = _AUTO_CREATE_MAPPING.get(key)
    attr = (
        config["attr"]
        if config
        else next(
            (
                field.name
                for field in related._meta.fields
                if field.name in {"name", "title", "number"}
            ),
            None,
        )
    )
    if not attr:
        return None
    defaults = config.get("defaults", {}) if config else {}
    obj, _ = related.objects.get_or_create(**{attr: text}, defaults=defaults)
    return obj.pk


_AUTO_CREATE_MAPPING = {
    "deals.insurancecompany": {"attr": "name"},
    "deals.insurancetype": {"attr": "name"},
    "deals.saleschannel": {"attr": "name"},
}


CREATED_POLICY_IDS: set[str] = set()
CREATED_PAYMENT_IDS: set[str] = set()
CREATED_DEAL_IDS: set[str] = set()


CLIENT_MAPPING_FILE = BASE_DIR / "import/data" / "client_mapping.json"


def _load_client_mapping() -> dict[str, str]:
    if not CLIENT_MAPPING_FILE.exists():
        return {}
    try:
        data = json.loads(CLIENT_MAPPING_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    return {str(key): str(value) for key, value in data.items()}


CLIENT_ID_OVERRIDES: dict[str, str] = _load_client_mapping()


def _normalize_legacy_key(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    key = str(value).strip()
    if not key:
        return None
    return key


def _map_to_client_uuid(value: Any) -> str | None:
    key = _normalize_legacy_key(value)
    if not key:
        return None
    return CLIENT_ID_OVERRIDES.get(key)


DEAL_ID_OVERRIDES: dict[str, str] = {}


def _map_to_deal_uuid(value: Any) -> str | None:
    key = _normalize_legacy_key(value)
    if not key:
        return None
    return DEAL_ID_OVERRIDES.get(key)


def _ensure_deal_uuid(value: Any) -> str | None:
    key = _normalize_legacy_key(value)
    if not key:
        return None
    mapped = DEAL_ID_OVERRIDES.get(key)
    if mapped:
        return mapped
    mapped = str(uuid.uuid4())
    DEAL_ID_OVERRIDES[key] = mapped
    return mapped


POLICY_ID_OVERRIDES: dict[str, str] = {}


def _map_to_policy_uuid(value: Any) -> str | None:
    key = _normalize_legacy_key(value)
    if not key:
        return None
    return POLICY_ID_OVERRIDES.get(key)


def _ensure_policy_uuid(value: Any) -> str | None:
    key = _normalize_legacy_key(value)
    if not key:
        return None
    mapped = POLICY_ID_OVERRIDES.get(key)
    if mapped:
        return mapped
    mapped = str(uuid.uuid4())
    POLICY_ID_OVERRIDES[key] = mapped
    return mapped


PAYMENT_ID_OVERRIDES: dict[str, str] = {}


def _map_to_payment_uuid(value: Any) -> str | None:
    key = _normalize_legacy_key(value)
    if not key:
        return None
    return PAYMENT_ID_OVERRIDES.get(key)


def _ensure_payment_uuid(value: Any) -> str | None:
    key = _normalize_legacy_key(value)
    if not key:
        return None
    mapped = PAYMENT_ID_OVERRIDES.get(key)
    if mapped:
        return mapped
    mapped = str(uuid.uuid4())
    PAYMENT_ID_OVERRIDES[key] = mapped
    return mapped


def _ensure_client_uuid_for_legacy(value: Any) -> str | None:
    key = _normalize_legacy_key(value)
    if not key:
        return None
    mapped = CLIENT_ID_OVERRIDES.get(key)
    if mapped:
        return mapped
    mapped = str(uuid.uuid4())
    CLIENT_ID_OVERRIDES[key] = mapped
    return mapped


def _parse_date(value: Any) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Не удалось разобрать дату: {value}")


def _parse_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    for fmt in ("%Y-%m-%d",):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ValueError(f"Не удалось разобрать дату/время: {value}")


def _deals_post_process(prepared: dict[str, Any], payload: Mapping[str, Any]) -> None:
    title = payload.get("calculations") or payload.get("description")
    prepared.setdefault(
        "title", str(title).strip() if title else f"Deal {payload.get('id') or 'row'}"
    )
    if payload.get("is_closed") and prepared.get("status") != "closed":
        prepared["status"] = "closed"
    if not prepared.get("loss_reason"):
        prepared["loss_reason"] = ""
    if prepared.get("title") and len(prepared["title"]) > 255:
        prepared["title"] = prepared["title"][:255]
    if not prepared.get("next_contact_date"):
        fallback = prepared.get("expected_close")
        if not fallback:
            try:
                fallback = _parse_date(payload.get("start_date"))
            except ValueError:
                fallback = None
        if fallback:
            prepared["next_contact_date"] = fallback


def _tasks_post_process(prepared: dict[str, Any], payload: Mapping[str, Any]) -> None:
    if "status" not in prepared:
        dispatch_state = str(payload.get("dispatch_state") or "").lower()
        mapping = {"idle": "todo", "in_progress": "in_progress", "done": "done"}
        prepared["status"] = mapping.get(dispatch_state, "todo")
    if payload.get("is_done"):
        prepared["status"] = "done"


def _ensure_sign(prepared: dict[str, Any], positive: bool) -> None:
    amount = prepared.get("amount")
    if amount is None:
        return
    decimal_amount = Decimal(amount)
    prepared["amount"] = decimal_amount if positive else -abs(decimal_amount)


SHEET_SPECS: Mapping[str, SheetSpec] = {
    "clients": SheetSpec(
        model_path="clients.Client",
        field_map={
            "name": "name",
            "phone": "phone",
            "note": "notes",
        },
        defaults={"notes": "", "phone": "+7 000 000 00 00"},
    ),
    "deals": SheetSpec(
        model_path="deals.Deal",
        field_map={
            "client": "client_id",
            "status": "status",
            "description": "description",
            "reminder_date": "next_contact_date",
            "start_date": "expected_close",
            "closed_reason": "loss_reason",
        },
        post_process=_deals_post_process,
    ),
    "policies": SheetSpec(
        model_path="policies.Policy",
        field_map={
            "client": "client_id",
            "deal": "deal_id",
            "policy_number": "number",
            "insurance_type": "insurance_type_id",
            "insurance_company": "insurance_company_id",
            "contractor": "counterparty",
            "sales_channel": "sales_channel_id",
            "start_date": "start_date",
            "end_date": "end_date",
            "vehicle_brand": "brand",
            "vehicle_model": "model",
            "vehicle_vin": "vin",
            "note": "note",
        },
    ),
    "payments": SheetSpec(
        model_path="finances.Payment",
        field_map={
            "policy": "policy_id",
            "amount": "amount",
            "payment_date": "scheduled_date",
            "actual_payment_date": "actual_date",
        },
    ),
    "incomes": SheetSpec(
        model_path="finances.FinancialRecord",
        field_map={
            "payment": "payment_id",
            "amount": "amount",
            "received_date": "date",
            "commission_source": "description",
            "note": "note",
        },
        defaults={"source": "income"},
        post_process=lambda prepared, _: _ensure_sign(prepared, positive=True),
    ),
    "expenses": SheetSpec(
        model_path="finances.FinancialRecord",
        field_map={
            "payment": "payment_id",
            "amount": "amount",
            "expense_date": "date",
            "expense_type": "description",
            "note": "note",
        },
        defaults={"source": "expense"},
        post_process=lambda prepared, _: _ensure_sign(prepared, positive=False),
    ),
    "tasks": SheetSpec(
        model_path="tasks.Task",
        field_map={
            "title": "title",
            "note": "description",
            "due_date": "due_at",
            "deal": "deal_id",
        },
        post_process=_tasks_post_process,
    ),
}


AUTO_CLEAR_SHEETS: set[str] = {"policies", "payments", "incomes", "expenses"}


def _truncate_model(model_path: str) -> None:
    model = apps.get_model(model_path)
    table = connection.ops.quote_name(model._meta.db_table)
    with connection.cursor() as cursor:
        cursor.execute(f"TRUNCATE TABLE {table} CASCADE;")


def _import_sheet(sheet: Worksheet, spec: SheetSpec, dry_run: bool) -> dict[str, int]:
    model = apps.get_model(spec.model_path)
    instances = []
    processed = 0

    for row_index, payload in _iter_rows(sheet):
        prepared: dict[str, Any] = dict(spec.defaults)
        for column, value in payload.items():
            target = spec.field_map.get(column)
            if not target:
                continue
            field_name = target[:-3] if target.endswith("_id") else target
            try:
                field_obj = model._meta.get_field(field_name)
            except FieldDoesNotExist:
                continue
            treat_as_id = target.endswith("_id") or isinstance(
                field_obj, models.ForeignKey
            )
            coerced = _coerce_value(field_obj, value, treat_as_id)
            if coerced is None and target in spec.defaults:
                continue
            prepared[target] = coerced

        if spec.post_process:
            spec.post_process(prepared, payload)

        if not prepared:
            continue

        if spec.model_path == "clients.Client":
            client_uuid = _ensure_client_uuid_for_legacy(payload.get("id"))
            if client_uuid:
                prepared["id"] = client_uuid
        elif spec.model_path == "deals.Deal":
            deal_uuid = _ensure_deal_uuid(payload.get("id"))
            if deal_uuid:
                prepared["id"] = deal_uuid
        elif spec.model_path == "policies.Policy":
            policy_uuid = _ensure_policy_uuid(payload.get("id"))
            if policy_uuid:
                prepared["id"] = policy_uuid
            if not prepared.get("deal_id"):
                print(f"skip policy row {row_index} without deal_id")
                continue
            if prepared["deal_id"] not in CREATED_DEAL_IDS:
                print(f"skip policy row {row_index} for missing deal {prepared['deal_id']}")
                continue
            if not prepared.get("insurance_type_id"):
                print(f"skip policy row {row_index} without insurance_type")
                continue
            if not prepared.get("insurance_company_id"):
                print(f"skip policy row {row_index} without insurance_company")
                continue
        elif spec.model_path == "finances.Payment":
            payment_uuid = _ensure_payment_uuid(payload.get("id"))
            if payment_uuid:
                prepared["id"] = payment_uuid

        if spec.model_path == "finances.Payment":
            policy_id = prepared.get("policy_id")
            if policy_id and policy_id not in CREATED_POLICY_IDS:
                print(f"skip payment row {row_index} for missing policy {policy_id}")
                continue
        if spec.model_path == "finances.FinancialRecord":
            payment_id = prepared.get("payment_id")
            if payment_id and payment_id not in CREATED_PAYMENT_IDS:
                print(
                    f"skip financial record row {row_index} for missing payment {payment_id}"
                )
                continue

        instances.append(model(**prepared))
        processed += 1

        if spec.model_path == "policies.Policy" and prepared.get("id"):
            CREATED_POLICY_IDS.add(prepared["id"])
        if spec.model_path == "finances.Payment" and prepared.get("id"):
            CREATED_PAYMENT_IDS.add(prepared["id"])
        if spec.model_path == "deals.Deal" and prepared.get("id"):
            CREATED_DEAL_IDS.add(prepared["id"])

    if not instances:
        return {"processed": 0, "created": 0}

    if not dry_run:
        print(f"bulk inserting {len(instances)} rows for {spec.model_path}")
        try:
            model.objects.bulk_create(instances)
        except DataError as exc:
            print("bulk insert failed", spec.model_path, exc)
            for instance in instances:
                try:
                    instance.save()
                    instance.delete()
                except DataError as inner:
                    print(
                        "single row failed", spec.model_path, instance.__dict__, inner
                    )
                    raise
            raise

    return {"processed": processed, "created": len(instances)}


def main() -> None:
    args = _parse_args()
    requested_path = Path(args.path)
    workbook_path = (
        requested_path if requested_path.is_absolute() else BASE_DIR / requested_path
    )
    workbook = load_workbook(filename=workbook_path, data_only=True)
    selected = _normalize_sheet_name(args.sheet) if args.sheet else None
    summary = []
    if args.clear:
        print("Очистка связанных таблиц перед импортом...")
        _clear_tables()

    cleared: set[str] = set()
    for sheet in workbook.worksheets:
        normalized = _normalize_sheet_name(sheet.title)
        if selected and normalized != selected:
            continue
        spec = SHEET_SPECS.get(normalized)
        if not spec:
            print(f"Пропускаю лист '{sheet.title}' — нет конфигурации.")
            continue
        if normalized in AUTO_CLEAR_SHEETS and normalized not in cleared:
            print(f"Clearing data for '{sheet.title}' before import")
            _truncate_model(spec.model_path)
            cleared.add(normalized)

        result = _import_sheet(sheet, spec, args.dry_run)
        summary.append((sheet.title, result))

    if not summary:
        print("Не найдено листов для импорта.")
        return

    for title, result in summary:
        print(
            f"{title}: обработано строк {result['processed']}, создано объектов {result['created']}"
        )

    if args.dry_run:
        print("Dry run — записи не выполнялись.")
    else:
        print("Импорт завершён.")


if __name__ == "__main__":
    main()
